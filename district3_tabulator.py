#!/usr/bin/env python3
"""
=============================================================================
3rd Congressional District Democratic Convention — May 2, 2026
ELECTION TABULATION TOOL  v3.0  —  Modular Ranked-Choice IRV Tabulator
=============================================================================
Handles three elections:
  A. State Central Committee (SCC)   — Sections V.K / VI    (IRV, 4 seats each)
  B. DEI Committee Chair             — Sections V.K / IX    (IRV, 1 seat)
  C. State Convention Committee      — Sections V.K / XII   (Slate vote, 14 seats)

Ballot input:
  • Google Sheets CSV export   (primary — electronic ballots)
  • Paper ballot CSV           (optional — merged by delegate number)

Other inputs:
  • surrendered_delegates.json  (snapshot from Credentials Tracker)
    Save this file in the same folder as the tabulator script.
    Generate the snapshot before the election begins.
    Supports two status types: "surrendered" and "non-issued".
    Revoked credentials should be processed as "surrendered" with a note.

Outputs:
  • Round-by-round results on screen
  • Excel workbook  (Summary + Redistribution Flow + one sheet per round)
  • Plain-text audit log  (for the official record)
  • Exceptions report  (Surrendered / Non-Issued / Exhausted)

Modules:
  • district3_irv.py     — core IRV/ranked-choice logic
  • district3_runoff.py  — runoff/tiebreak ballot logic

Requirements:  Python 3.8+,  openpyxl  (pip install openpyxl)
=============================================================================

SCC SEAT-RESTART RULE (confirmed with IDP contact 2026-04-28):
  After each SCC seat is won, ONLY the winner is removed from the pool.
  All other candidates — including those eliminated during that seat's
  rounds — return to the pool for the next seat.  Example:
    Seat 1: A wins, C & D eliminated → Seat 2 pool: B, C, D, E, F, ...
    Seat 2: B wins, E eliminated     → Seat 3 pool: C, D, E, F, ...
  Each seat starts a fresh IRV cycle with the full non-elected pool.
=============================================================================
"""

import csv
import json
import os
import re
import sys
import datetime
from collections import defaultdict
from pathlib import Path

# ── Try to import openpyxl ──────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── Import tabulation modules ──────────────────────────────────────────────
from district3_irv import (
    count_round_votes,
    compute_threshold_elimination,
    drop_lowest_candidate,
    run_irv_one_seat,
    init_irv_display,
)
from district3_runoff import (
    run_runoff,
    init_runoff_display,
)

# =============================================================================
# CONSTANTS
# =============================================================================

VERSION           = "3.0.0"
CONVENTION_DATE   = "May 2, 2026"
DISTRICT          = "3rd Congressional District Democratic Convention"

ELIMINATION_THRESHOLD   = 0.15   # 15%  — candidates below this are eliminated
SCC_FIRST_BALLOT_MAX    = 0.50   # 50%  — max fraction of seats filled in Round 1
MAJORITY_THRESHOLD      = 0.50   # must receive MORE THAN 50% to be elected

# Column header patterns in the Google Sheets CSV export
RANK_HEADER_RE = re.compile(r'candidate\s+([A-Za-z])\s+rank', re.IGNORECASE)

# Rank-position header pattern for test/paper ballots (1st, 2nd, 3rd, 4th, …)
RANK_POSITION_RE = re.compile(r'^(\d+)(?:st|nd|rd|th)$', re.IGNORECASE)

# =============================================================================
# COLOR / DISPLAY HELPERS
# =============================================================================

class Color:
    RESET   = "\033[0m"
    BOLD    = "\033[1m"
    GREEN   = "\033[92m"
    YELLOW  = "\033[93m"
    RED     = "\033[91m"
    CYAN    = "\033[96m"
    MAGENTA = "\033[95m"
    WHITE   = "\033[97m"
    GREY    = "\033[90m"

def c(text, color):
    return f"{color}{text}{Color.RESET}"

def banner():
    w = 70
    print()
    print(c("=" * w, Color.CYAN))
    print(c(f"  {DISTRICT}".center(w), Color.BOLD + Color.WHITE))
    print(c(f"  Election Tabulation Tool  v{VERSION}".center(w), Color.CYAN))
    print(c(f"  {CONVENTION_DATE}".center(w), Color.GREY))
    print(c("=" * w, Color.CYAN))
    print()

def section_header(title):
    print()
    print(c("─" * 62, Color.CYAN))
    print(c(f"  {title}", Color.BOLD + Color.WHITE))
    print(c("─" * 62, Color.CYAN))
    log(f"\n{'─'*62}\n  {title}\n{'─'*62}", also_print=False)

def seat_ballot_header(seat_num, election_label):
    msg = f"  SEAT {seat_num} BALLOT — {election_label}"
    print()
    print(c("╔" + "═" * 60 + "╗", Color.MAGENTA + Color.BOLD))
    print(c(f"║{msg:<60}║", Color.MAGENTA + Color.BOLD))
    print(c("╚" + "═" * 60 + "╝", Color.MAGENTA + Color.BOLD))
    log(f"\n{'═'*62}\n[SEAT {seat_num} BALLOT] {election_label}\n{'═'*62}",
        also_print=False)

def distribution_round_header(seat_round_num, seat_num, election_label):
    msg = f"  Seat {seat_num} Ballot — Distribution Round {seat_round_num}"
    print()
    print(c("┌" + "─" * 60 + "┐", Color.YELLOW))
    print(c(f"│{msg:<60}│", Color.YELLOW))
    print(c("└" + "─" * 60 + "┘", Color.YELLOW))
    log(f"\n[SEAT {seat_num} — Distribution Round {seat_round_num}] "
        f"{election_label}", also_print=False)

def display_round_table(candidate_map, vote_counts, total_active,
                        majority_needed, elected=None, eliminated=None):
    elected    = elected    or []
    eliminated = eliminated or []
    threshold_votes = total_active * ELIMINATION_THRESHOLD
    sorted_cands = sorted(vote_counts.items(), key=lambda x: -x[1])
    col_w = [32, 8, 10, 16]
    hdr = f"  {'Candidate':<{col_w[0]}} {'Votes':>{col_w[1]}} {'%':>{col_w[2]}} {'Status':>{col_w[3]}}"
    print(c("\n" + hdr, Color.BOLD))
    print(c("  " + "─" * sum(col_w), Color.GREY))
    for num, votes in sorted_cands:
        name = candidate_map.get(num, f"Candidate {num}")
        pct = votes / total_active * 100 if total_active > 0 else 0
        if num in elected:
            status = "✓ ELECTED"
            clr = Color.GREEN + Color.BOLD
        elif num in eliminated:
            status = "✗ ELIMINATED"
            clr = Color.RED
        elif votes < threshold_votes:
            status = "below 15%"
            clr = Color.YELLOW
        else:
            status = ""
            clr = Color.WHITE
        line = f"  {name:<{col_w[0]}} {votes:>{col_w[1]}} {pct:>{col_w[2]-1}.1f}% {status:>{col_w[3]}}"
        print(c(line, clr))
    print(c(f"\n  Active ballots: {total_active}   "
            f"Majority needed: {majority_needed}   "
            f"15% threshold: {threshold_votes:.1f} votes", Color.GREY))

# ── Prompts ─────────────────────────────────────────────────────────────────

def prompt(msg, default=None):
    result = input(msg).strip()
    return result if result else default

def confirm(msg):
    resp = input(f"{msg} (y/n): ").strip().lower()
    return resp in ("y", "yes")

def menu(title, options):
    print(c(f"\n  {title}", Color.BOLD))
    for i, opt in enumerate(options, 1):
        print(c(f"    {i}. {opt}", Color.WHITE))
    while True:
        try:
            choice = int(input(c("\n  Enter choice: ", Color.CYAN)).strip())
            if 1 <= choice <= len(options):
                return choice
        except ValueError:
            pass
        print(c("  Please enter a valid number.", Color.RED))

# =============================================================================
# AUDIT LOG
# =============================================================================

_audit_lines = []

def log(text, also_print=True):
    _audit_lines.append(text)
    if also_print:
        print(text)

def save_audit_log(filepath):
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(_audit_lines))
    print(c(f"\n  Audit log saved → {filepath}", Color.GREEN))

# =============================================================================
# INITIALIZE MODULE DISPLAY FUNCTIONS
# =============================================================================
# Inject display/logging into the IRV and runoff modules so they can print
# formatted output without circular imports.

init_irv_display(c, log, seat_ballot_header, distribution_round_header,
                 display_round_table, Color)
init_runoff_display(c, log, section_header, Color, None, confirm)
# Note: get_csv_filepath is set below after it's defined.

# =============================================================================
# BALLOT DATA CLASS
# =============================================================================

class Ballot:
    """Represents one delegate's ranked-choice ballot."""
    __slots__ = ("delegate_number", "timestamp", "rankings",
                 "source", "is_spoiled", "spoil_reason")

    def __init__(self, delegate_number, timestamp, rankings,
                 source="electronic"):
        self.delegate_number = str(delegate_number).strip()
        self.timestamp       = timestamp
        self.rankings        = rankings
        self.source          = source
        self.is_spoiled      = False
        self.spoil_reason    = ""

    def get_first_active_choice(self, active_keys):
        best_num  = None
        best_rank = None
        for num in active_keys:
            rank = self.rankings.get(num)
            if rank and isinstance(rank, int) and rank > 0:
                if best_rank is None or rank < best_rank:
                    best_rank = rank
                    best_num  = num
        return best_num

# =============================================================================
# CSV LOADING — ELECTRONIC BALLOTS (Google Sheets export)
# =============================================================================

def load_electronic_ballots(filepath):
    ballots        = []
    warnings       = []
    candidate_keys = []

    try:
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader     = csv.DictReader(f)
            fieldnames = [fn.strip() for fn in (reader.fieldnames or [])]

            # --- FORMAT DETECTION ---
            # Format A: "Candidate A Rank" columns (Google Sheets export)
            rank_cols = {}
            for fn in fieldnames:
                m = RANK_HEADER_RE.match(fn.strip())
                if m:
                    rank_cols[m.group(1).upper()] = fn

            # Format B: Rank-position columns (1st, 2nd, 3rd, …)
            position_cols = {}          # {1: "1st", 2: "2nd", ...}
            for fn in fieldnames:
                m = RANK_POSITION_RE.match(fn.strip())
                if m:
                    position_cols[int(m.group(1))] = fn

            use_position_format = False
            if rank_cols:
                # Prefer Format A when present
                candidate_keys = sorted(rank_cols.keys())
            elif position_cols:
                # Use Format B — candidate letters appear as cell values
                use_position_format = True
            else:
                warnings.append("No 'Candidate A Rank' or rank-position "
                                "(1st, 2nd, …) columns found.")
                return [], [], warnings

            # For position format, we discover candidate keys on first pass
            discovered_candidates = set()

            for row_num, row in enumerate(reader, start=2):
                delegate_num = str(row.get("Delegate Number",
                                  row.get("Delegate_Number",
                                  row.get("delegate_number", "")))).strip()
                timestamp    = str(row.get("Timestamp", "")).strip()
                is_test      = str(row.get("Is Test", "NO")).strip().upper()

                if is_test == "YES":
                    continue

                if use_position_format:
                    # Read candidate letters from each rank column
                    rankings = {}
                    for pos, col in sorted(position_cols.items()):
                        cand_letter = str(row.get(col, "")).strip().upper()
                        if cand_letter and len(cand_letter) == 1 and cand_letter.isalpha():
                            rankings[cand_letter] = pos
                            discovered_candidates.add(cand_letter)
                        # Blank = unranked at this position, skip
                else:
                    # Original format: rank numbers in candidate columns
                    rankings = {}
                    for cand_num, col in rank_cols.items():
                        raw = str(row.get(col, "")).strip()
                        if raw == "":
                            rankings[cand_num] = 0
                        else:
                            try:
                                rankings[cand_num] = int(raw)
                            except ValueError:
                                rankings[cand_num] = 0
                                warnings.append(
                                    f"Row {row_num}: non-integer rank '{raw}' "
                                    f"for Candidate {cand_num}")

                if not delegate_num:
                    warnings.append(
                        f"Row {row_num}: missing delegate number — skipped")
                    continue

                ballots.append(Ballot(delegate_num, timestamp, rankings,
                                      source="electronic"))

            # Set candidate_keys for position format after reading all rows
            if use_position_format:
                candidate_keys = sorted(discovered_candidates)

    except FileNotFoundError:
        warnings.append(f"File not found: {filepath}")

    return ballots, candidate_keys, warnings

# =============================================================================
# CSV LOADING — PAPER BALLOTS
# =============================================================================

def load_paper_ballots(filepath):
    ballots, candidate_keys, warnings = load_electronic_ballots(filepath)
    for b in ballots:
        b.source = "paper"
    return ballots, candidate_keys, warnings

# =============================================================================
# BALLOT MERGE  (Paper-first with ranking fallback)
# =============================================================================

def _count_rankings(ballot):
    """Count how many candidates a ballot actually ranked (rank > 0)."""
    return sum(1 for r in ballot.rankings.values()
               if isinstance(r, int) and r > 0)

def merge_ballots(electronic, paper):
    """
    Merge electronic and paper ballots.
    Priority: PAPER over electronic, UNLESS the paper ballot has fewer
    rankings than the electronic one.  In that case, whichever ballot
    holds the most rankings is accepted.
    """
    by_delegate = {}

    electronic_by_dn = {}
    for b in electronic:
        electronic_by_dn[b.delegate_number] = b

    paper_by_dn = {}
    for b in paper:
        paper_by_dn[b.delegate_number] = b

    all_delegates = set(electronic_by_dn.keys()) | set(paper_by_dn.keys())

    paper_used = 0
    electronic_used = 0
    conflicts = 0

    for dn in all_delegates:
        e_ballot = electronic_by_dn.get(dn)
        p_ballot = paper_by_dn.get(dn)

        if p_ballot and not e_ballot:
            by_delegate[dn] = p_ballot
            paper_used += 1
        elif e_ballot and not p_ballot:
            by_delegate[dn] = e_ballot
            electronic_used += 1
        else:
            conflicts += 1
            p_ranks = _count_rankings(p_ballot)
            e_ranks = _count_rankings(e_ballot)
            if p_ranks >= e_ranks:
                by_delegate[dn] = p_ballot
                paper_used += 1
                log(f"  [MERGE] Delegate {dn}: paper ballot used "
                    f"({p_ranks} rankings vs {e_ranks} electronic).",
                    also_print=False)
            else:
                by_delegate[dn] = e_ballot
                electronic_used += 1
                log(f"  [MERGE] Delegate {dn}: electronic ballot used "
                    f"— paper had fewer rankings "
                    f"({p_ranks} vs {e_ranks}).", also_print=False)

    merged = list(by_delegate.values())
    log(f"\n[MERGE] Electronic: {len(electronic)}  Paper: {len(paper)}  "
        f"Conflicts: {conflicts}  Total: {len(merged)}", also_print=False)
    return merged

# =============================================================================
# SURRENDERED CREDENTIALS CHECK
# =============================================================================

def _parse_timestamp(ts_str):
    """
    Parse a timestamp string from either:
      - ISO format:       "2026-05-02T09:30:00" or "2026-05-02 09:30:00"
      - toLocaleString(): "5/2/2026, 9:35:00 AM"  (US locale, Apps Script)
    Returns a datetime or None if unparseable.
    """
    if not ts_str or not isinstance(ts_str, str):
        return None
    ts_str = ts_str.strip()

    try:
        return datetime.datetime.fromisoformat(ts_str)
    except ValueError:
        pass

    for fmt in ("%m/%d/%Y, %I:%M:%S %p",
                "%m/%d/%Y %I:%M:%S %p",
                "%m/%d/%Y, %H:%M:%S",
                "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.datetime.strptime(ts_str, fmt)
        except ValueError:
            continue

    return None


def check_surrendered(ballots, surrendered_path):
    """
    Load surrendered_delegates.json and flag any ballot from a delegate
    whose credentials are invalid (surrendered or non-issued).

    surrendered_delegates.json format:
      [
        { "delegate_number": "101", "name": "Jane Doe",
          "status": "surrendered",
          "surrender_time": "2026-05-02 09:30:00" },
        { "delegate_number": "205", "name": "John Smith",
          "status": "non-issued",
          "note": "Never checked in" }
      ]

    Supported status values:
      - "surrendered" — delegate checked in then surrendered credentials
      - "non-issued"  — delegate was on the roster but never checked in
      - (any other)   — treated as surrendered for flagging purposes

    If the "status" field is missing, defaults to "surrendered" for
    backward compatibility.

    Revoked credentials: process as "surrendered" in the Credentials
    Tracker and add the reason in the "note" field.

    Place this file in the same directory as the tabulator script.
    Generate the snapshot from the Credentials Tracker Google Sheet
    before the election begins.

    Returns (valid_ballots, spoiled_credentials) lists.
    """
    flagged_set  = set()
    flagged_info = {}

    if not os.path.isfile(surrendered_path):
        log("  [CREDENTIALS] No surrendered_delegates.json found — "
            "skipping check.", also_print=False)
        print(c(f"  ℹ  No surrendered credentials file found at:",
                Color.GREY))
        print(c(f"      {surrendered_path}", Color.GREY))
        print(c(f"      (This is normal if no delegates surrendered.)",
                Color.GREY))
        return ballots, []

    try:
        with open(surrendered_path, encoding="utf-8") as f:
            records = json.load(f)
        for rec in records:
            dn = str(rec.get("delegate_number", "")).strip()
            if dn:
                flagged_set.add(dn)
                flagged_info[dn] = {
                    "name":   rec.get("name", "Unknown"),
                    "status": rec.get("status", "surrendered").strip().lower(),
                    "time":   rec.get("surrender_time", ""),
                    "note":   rec.get("note", ""),
                }
    except (json.JSONDecodeError, OSError) as e:
        log(f"  [WARNING] Could not read surrendered_delegates.json: {e}",
            also_print=False)
        return ballots, []

    if not flagged_set:
        log("  [CREDENTIALS] File loaded but no entries found.",
            also_print=False)
        return ballots, []

    n_surrendered = sum(1 for i in flagged_info.values()
                        if i["status"] != "non-issued")
    n_nonissued   = sum(1 for i in flagged_info.values()
                        if i["status"] == "non-issued")
    log(f"  [CREDENTIALS] Loaded {len(flagged_set)} flagged delegate(s) "
        f"({n_surrendered} surrendered, {n_nonissued} non-issued).",
        also_print=False)

    valid   = []
    spoiled = []

    for b in ballots:
        if b.delegate_number in flagged_set:
            info   = flagged_info[b.delegate_number]
            status = info["status"]

            if status == "non-issued":
                # Delegate never checked in — ballot should not exist
                note_detail = f" — {info['note']}" if info["note"] else ""
                b.is_spoiled   = True
                b.spoil_reason = f"Non-Issued Credentials{note_detail}"
            else:
                # Surrendered (or revoked processed as surrendered)
                surrender_ts = info["time"]
                ballot_dt    = _parse_timestamp(b.timestamp)
                surrender_dt = _parse_timestamp(surrender_ts)

                # Only spoil if the delegate surrendered BEFORE voting.
                # If they voted first and surrendered later, the vote
                # was cast with valid credentials and should count.
                if ballot_dt and surrender_dt and ballot_dt < surrender_dt:
                    # Voted before surrendering — ballot is valid
                    valid.append(b)
                    log(f"  [VALID] Delegate {b.delegate_number} "
                        f"({info['name']}) voted {ballot_dt.strftime('%H:%M')}"
                        f" before surrendering {surrender_dt.strftime('%H:%M')}"
                        f" — ballot counts.", also_print=False)
                    continue

                time_detail = ""
                if ballot_dt and surrender_dt:
                    time_detail = (
                        f" (surrendered {surrender_dt.strftime('%H:%M')}, "
                        f"voted {ballot_dt.strftime('%H:%M')})")
                elif surrender_dt:
                    time_detail = (
                        f" (surrendered {surrender_dt.strftime('%H:%M')})")

                note_detail = f" — {info['note']}" if info["note"] else ""
                b.is_spoiled   = True
                b.spoil_reason = (f"Surrendered Credentials"
                                  f"{time_detail}{note_detail}")

            spoiled.append(b)
            log(f"  [SPOILED] Delegate {b.delegate_number} "
                f"({info['name']}) — {b.spoil_reason}", also_print=False)
        else:
            valid.append(b)

    return valid, spoiled

# =============================================================================
# CANDIDATE SETUP
# =============================================================================

def enter_candidates(election_name, found_keys=None):
    print()
    if found_keys:
        print(c(f"  Found {len(found_keys)} candidate column(s) in the CSV: "
                f"{', '.join(found_keys)}", Color.CYAN))
    print(c(f"  Enter candidate names for: {election_name}", Color.BOLD))
    print(c("  (Type 'done' when finished, or press Enter to accept "
            "letter-only labels.)\n", Color.GREY))

    candidate_map = {}
    if found_keys:
        for key in found_keys:
            name = input(c(f"    Candidate {key}: ", Color.WHITE)).strip()
            if name.lower() == "done":
                break
            candidate_map[key] = name if name else f"Candidate {key}"
    else:
        letter = ord("A")
        while True:
            key = chr(letter)
            name = input(c(f"    Candidate {key}: ", Color.WHITE)).strip()
            if name.lower() == "done" or name == "":
                break
            candidate_map[key] = name
            letter += 1

    log(f"[CANDIDATES] {election_name}: {candidate_map}", also_print=False)
    return candidate_map

def get_csv_filepath(label):
    print(c(f"\n  Load CSV for: {label}", Color.BOLD))
    path = input(c("  File path (or drag & drop): ", Color.CYAN)).strip()
    path = path.strip('"').strip("'")
    if path and os.path.isfile(path):
        return path
    elif path:
        print(c(f"  File not found: {path}", Color.RED))
    return None

def load_ballots_for_election(election_name, data_dir):
    section_header(f"BALLOT LOADING — {election_name.upper()}")

    electronic_ballots = []
    paper_ballots      = []
    candidate_keys     = []

    choice = menu(f"How will {election_name} ballots be loaded?", [
        "Google Sheets CSV export  (electronic ballots)",
        "Paper ballot CSV only",
        "Both electronic + paper  (will merge by delegate number)",
    ])

    if choice in (1, 3):
        path = get_csv_filepath(f"{election_name} — Electronic")
        if path:
            electronic_ballots, candidate_keys, warnings = (
                load_electronic_ballots(path))
            for w in warnings:
                print(c(f"  ⚠  {w}", Color.YELLOW))
            print(c(f"  Loaded {len(electronic_ballots)} electronic ballot(s) "
                    f"with {len(candidate_keys)} candidate column(s).",
                    Color.GREEN))

    if choice in (2, 3):
        path = get_csv_filepath(f"{election_name} — Paper")
        if path:
            paper_ballots, paper_keys, warnings = load_paper_ballots(path)
            for w in warnings:
                print(c(f"  ⚠  {w}", Color.YELLOW))
            print(c(f"  Loaded {len(paper_ballots)} paper ballot(s).",
                    Color.GREEN))
            if not candidate_keys:
                candidate_keys = paper_keys

    if not electronic_ballots and not paper_ballots:
        print(c("\n  No ballots loaded — cannot tabulate.", Color.RED))
        return None, None, None

    all_ballots = merge_ballots(electronic_ballots, paper_ballots)

    surrendered_path = os.path.join(data_dir, "surrendered_delegates.json")
    all_ballots, spoiled_surrendered = check_surrendered(
        all_ballots, surrendered_path)
    if spoiled_surrendered:
        n_surr = sum(1 for b in spoiled_surrendered
                     if "Non-Issued" not in (b.spoil_reason or ""))
        n_ni   = sum(1 for b in spoiled_surrendered
                     if "Non-Issued" in (b.spoil_reason or ""))
        parts = []
        if n_surr:
            parts.append(f"{n_surr} Surrendered")
        if n_ni:
            parts.append(f"{n_ni} Non-Issued")
        print(c(f"\n  ⚠  {len(spoiled_surrendered)} ballot(s) flagged: "
                f"{', '.join(parts)}.", Color.YELLOW))
        for b in spoiled_surrendered:
            clr = Color.RED if "Non-Issued" in (b.spoil_reason or "") else Color.YELLOW
            print(c(f"     Delegate {b.delegate_number} — {b.spoil_reason}",
                    clr))

    candidate_map = enter_candidates(election_name, candidate_keys)

    return all_ballots, candidate_map, spoiled_surrendered


# ── Now set get_csv_filepath in the runoff module ───────────────────────────
import district3_runoff as _runoff_mod
_runoff_mod._get_csv_filepath = get_csv_filepath

# =============================================================================
# ELECTION A — STATE CENTRAL COMMITTEE (SCC)   Article VI / V.K
# =============================================================================

def run_scc_election(gender_label, seats, ballots, candidate_map,
                     spoiled_surrendered, output_dir):
    """
    Run one gender category of the SCC election (sequential per-seat IRV).

    SEAT RESTART LOGIC (confirmed with IDP contact 2026-04-28):
      After each winner, only the winner is removed from the pool.
      All other candidates return — eliminations do NOT carry forward.
    """
    section_header(f"SCC — {gender_label.upper()} ({seats} seats)")
    log(f"\n[ELECTION] SCC — {gender_label} — {seats} seats",
        also_print=False)
    log(f"[DATE] {datetime.datetime.now().isoformat()}", also_print=False)

    all_nums    = list(candidate_map.keys())
    elected_all = []
    rounds_data = []

    election_key = ("scc-w" if "women" in gender_label.lower()
                    else "scc-mnb")

    max_seat1_round1_wins = max(1, int(seats * SCC_FIRST_BALLOT_MAX))
    print(c(f"\n  Seat 1 Ballot rule: no more than "
            f"{max_seat1_round1_wins} of {seats} seats may be filled "
            f"in the first distribution round.", Color.CYAN))
    log(f"[RULE] Seat 1 first-distribution-round max wins: "
        f"{max_seat1_round1_wins}", also_print=False)

    # Ratification path: nominees == seats
    if len(all_nums) == seats:
        print(c(f"\n  {len(all_nums)} nominees for {seats} seats — "
                f"proceeding to ratification ballot.", Color.YELLOW))
        vote_counts, total_active, exhausted = count_round_votes(
            ballots, all_nums)
        majority_needed = int(total_active * MAJORITY_THRESHOLD) + 1
        elected = [num for num, v in vote_counts.items()
                   if v >= majority_needed]
        display_round_table(candidate_map, vote_counts, total_active,
                            majority_needed, elected=elected)
        rounds_data.append({
            "round_num": 1, "seat_num": 0,
            "vote_counts": vote_counts, "total_active": total_active,
            "exhausted": exhausted, "majority_needed": majority_needed,
            "elected": elected, "eliminated": [],
        })
        log(f"[RESULT] Ratification — Elected: "
            f"{[candidate_map.get(n,n) for n in elected]}",
            also_print=False)
        if rounds_data:
            _save_scc_excel(gender_label, candidate_map, rounds_data,
                            output_dir)
        return elected, rounds_data, spoiled_surrendered

    # Sequential per-seat IRV
    for seat in range(1, seats + 1):
        if not candidate_map:
            print(c("\n  No candidates remaining.", Color.YELLOW))
            break

        active_pool = [num for num in all_nums
                       if num not in elected_all]

        if len(active_pool) == 0:
            print(c(f"\n  No remaining candidates for Seat {seat}.",
                    Color.YELLOW))
            break

        if len(active_pool) == 1:
            # Uncontested — still run the count for the official record
            winner_num = active_pool[0]
            name = candidate_map.get(winner_num, winner_num)

            seat_ballot_header(seat, f"SCC {gender_label}")
            distribution_round_header(1, seat, f"SCC {gender_label}")

            vote_counts, total_active, exhausted = count_round_votes(
                ballots, active_pool)
            majority_needed = int(total_active * MAJORITY_THRESHOLD) + 1
            winner_votes = vote_counts.get(winner_num, 0)
            winner_pct = (winner_votes / total_active * 100
                          if total_active > 0 else 0)

            display_round_table(candidate_map, vote_counts, total_active,
                                majority_needed, elected=[winner_num])
            rounds_data.append({
                "seat_num":        seat,
                "seat_round_num":  1,
                "vote_counts":     dict(vote_counts),
                "total_active":    total_active,
                "exhausted":       exhausted,
                "majority_needed": majority_needed,
                "elected":         [winner_num],
                "eliminated":      [],
                "elim_reason":     "uncontested — no candidates dropped",
            })

            elected_all.append(winner_num)
            print(c(f"\n  Uncontested — no candidates dropped.",
                    Color.YELLOW))
            print(c(f"  {name} received {winner_votes} votes "
                    f"({winner_pct:.1f}% of active ballots).",
                    Color.CYAN))
            print(c(f"  ✓ {name} ELECTED (Seat {seat})",
                    Color.GREEN + Color.BOLD))
            log(f"[SEAT {seat}] Uncontested: {name} — {winner_votes} "
                f"votes ({winner_pct:.1f}%)", also_print=False)
            continue

        winner_num = run_irv_one_seat(
            ballots, candidate_map, active_pool,
            f"SCC {gender_label}", seat,
            max_seat1_round1_wins if seat == 1 else 1,
            rounds_data,
            election_key=election_key,
        )

        if winner_num:
            elected_all.append(winner_num)
            print(c(f"  (All non-elected candidates return to pool "
                    f"for next seat.)", Color.GREY))
        else:
            print(c(f"\n  WARNING: No winner found for Seat {seat}.",
                    Color.RED))

    # Final results
    print()
    print(c(f"  FINAL RESULT — SCC {gender_label} ({seats} seats)",
            Color.BOLD + Color.WHITE))
    for i, num in enumerate(elected_all, 1):
        print(c(f"    {i}. {candidate_map.get(num, num)}",
                Color.GREEN + Color.BOLD))
    log(f"\n[FINAL] SCC {gender_label} — Elected: "
        f"{[candidate_map.get(n,n) for n in elected_all]}",
        also_print=False)

    _print_exceptions(ballots, spoiled_surrendered,
                      election_name=f"SCC {gender_label}")

    if rounds_data:
        _save_scc_excel(gender_label, candidate_map, rounds_data,
                        output_dir)

    return elected_all, rounds_data, spoiled_surrendered


def _save_scc_excel(gender_label, candidate_map, rounds_data, output_dir):
    safe = gender_label.replace("/", "_").replace(" ", "_")
    path = os.path.join(output_dir, f"SCC_{safe}_Results.xlsx")
    create_excel_report(f"SCC — {gender_label}", candidate_map,
                        rounds_data, path)


def run_scc(ballots_women, candidate_map_women, spoiled_women,
            ballots_mnb, candidate_map_mnb, spoiled_mnb, output_dir):
    """Run the full SCC election (Women + Men/Non-Binary)."""
    section_header("STATE CENTRAL COMMITTEE (SCC) ELECTION — Article VI")
    print(c("\n  Two separate ballots: Women (4 seats) and "
            "Men/Non-Binary (4 seats).\n", Color.CYAN))
    results = {}
    elected_w, _, _ = run_scc_election(
        "Women", 4, ballots_women, candidate_map_women,
        spoiled_women, output_dir)
    results["Women"] = elected_w
    if confirm("\n  Proceed to Men/Non-Binary ballot?"):
        elected_mnb, _, _ = run_scc_election(
            "Men_NonBinary", 4, ballots_mnb, candidate_map_mnb,
            spoiled_mnb, output_dir)
        results["Men_NonBinary"] = elected_mnb
    return results


# =============================================================================
# ELECTION B — DEI COMMITTEE CHAIR   Article IX / V.K
# =============================================================================

def run_dei_election(ballots, candidate_map, spoiled_surrendered,
                     output_dir):
    """
    Run the DEI Committee Chair election (1 seat, IRV with special tie rule).

    Two-step Round 1 (same as SCC):
      Step A: 15% threshold → redistribute → winner check
      Step B: Drop lowest (with runoff ballot for ties)

    Tie rule (Convention Rules V.K):
      If tied candidates share the lowest total and eliminating all
      would leave < 2, hold a special tiebreak RUNOFF BALLOT
      (not a recount of existing rankings).
    """
    section_header("DEI COMMITTEE CHAIR ELECTION — Article IX")
    log("\n[ELECTION] DEI Committee Chair — 1 seat", also_print=False)
    log(f"[DATE] {datetime.datetime.now().isoformat()}", also_print=False)

    all_nums    = list(candidate_map.keys())
    active      = list(all_nums)
    rounds_data = []
    seat_round  = 1
    winner_num  = None

    seat_ballot_header(1, "DEI Committee Chair")

    while not winner_num and active:
        distribution_round_header(seat_round, 1, "DEI Committee Chair")

        vote_counts, total_active, exhausted = count_round_votes(
            ballots, active)
        majority_needed = int(total_active * MAJORITY_THRESHOLD) + 1

        # Check for winner
        best_num = max(vote_counts, key=vote_counts.get)
        if vote_counts[best_num] >= majority_needed:
            winner_num = best_num
            display_round_table(candidate_map, vote_counts, total_active,
                                majority_needed, elected=[winner_num])
            rounds_data.append({
                "seat_num": 1, "seat_round_num": seat_round,
                "vote_counts": dict(vote_counts),
                "total_active": total_active,
                "exhausted": exhausted,
                "majority_needed": majority_needed,
                "elected": [winner_num], "eliminated": [],
            })
            name = candidate_map.get(winner_num, winner_num)
            print(c(f"\n  ✓ {name} ELECTED as DEI Committee Chair",
                    Color.GREEN + Color.BOLD))
            log(f"[SEAT 1 — Round {seat_round}] WINNER: {name}",
                also_print=False)
            break

        # ── No winner — elimination ────────────────────────────────────
        sorted_cands = sorted(vote_counts.items(), key=lambda x: x[1])
        to_eliminate = []
        elim_reason  = ""

        if seat_round == 1:
            # ── Step A: 15% threshold ──────────────────────────────────
            threshold_elim, eff_pct = compute_threshold_elimination(
                vote_counts)

            # DEI safeguard: guarantee >= 2 survivors
            survivors = [n for n in active if n not in threshold_elim]
            while len(survivors) < 2 and len(threshold_elim) > 1:
                by_votes = sorted(
                    [(n, vote_counts[n]) for n in threshold_elim],
                    key=lambda x: x[1])
                threshold_elim = [n for n, _ in by_votes[:-1]]
                survivors = [n for n in active
                             if n not in threshold_elim]

            if threshold_elim:
                elim_reason = (f"below {eff_pct:.1f}% "
                               f"elimination threshold")
                display_round_table(
                    candidate_map, vote_counts, total_active,
                    majority_needed, eliminated=threshold_elim)
                rounds_data.append({
                    "seat_num": 1, "seat_round_num": seat_round,
                    "vote_counts": dict(vote_counts),
                    "total_active": total_active,
                    "exhausted": exhausted,
                    "majority_needed": majority_needed,
                    "elected": [],
                    "eliminated": list(threshold_elim),
                    "elim_reason": elim_reason,
                })
                for num in threshold_elim:
                    name = candidate_map.get(num, num)
                    active.remove(num)
                    print(c(f"  ✗ {name} eliminated  ({elim_reason})",
                            Color.RED))

                if len(active) == 1:
                    winner_num = active[0]
                    name = candidate_map.get(winner_num, winner_num)
                    print(c(f"\n  Only one candidate remaining "
                            f"— {name} is elected.", Color.YELLOW))
                    print(c(f"  ✓ {name} ELECTED as DEI Committee Chair",
                            Color.GREEN + Color.BOLD))
                    seat_round += 1
                    continue

                # Recount after threshold redistribution
                vote_counts, total_active, exhausted = count_round_votes(
                    ballots, active)
                majority_needed = int(
                    total_active * MAJORITY_THRESHOLD) + 1

                best_num = max(vote_counts, key=vote_counts.get)
                if vote_counts[best_num] >= majority_needed:
                    winner_num = best_num
                    display_round_table(
                        candidate_map, vote_counts, total_active,
                        majority_needed, elected=[winner_num])
                    rounds_data.append({
                        "seat_num": 1, "seat_round_num": seat_round,
                        "vote_counts": dict(vote_counts),
                        "total_active": total_active,
                        "exhausted": exhausted,
                        "majority_needed": majority_needed,
                        "elected": [winner_num], "eliminated": [],
                    })
                    name = candidate_map.get(winner_num, winner_num)
                    print(c(f"\n  ✓ {name} ELECTED as DEI Committee "
                            f"Chair — after threshold redistribution",
                            Color.GREEN + Color.BOLD))
                    break

            # ── Step B: Drop lowest ────────────────────────────────────
            tied_lowest = drop_lowest_candidate(vote_counts)

        else:
            # ── Round 2+: lowest drops ─────────────────────────────────
            lowest_votes = sorted_cands[0][1]
            tied_lowest = [num for num, v in vote_counts.items()
                           if v == lowest_votes]

        # Handle the drop (shared by Round 1 Step B and Round 2+)
        if not winner_num and tied_lowest:
            if len(tied_lowest) > 1:
                after_elim = len(active) - len(tied_lowest)
                if after_elim >= 2:
                    to_eliminate = tied_lowest
                    names = [candidate_map.get(n, n)
                             for n in tied_lowest]
                    print(c(f"\n  Tie for lowest "
                            f"({', '.join(names)}) — all eliminated.",
                            Color.YELLOW))
                    elim_reason = "tied for lowest vote-getter"
                else:
                    # Tiebreak runoff ballot (Convention Rules V.K)
                    names = [candidate_map.get(n, n)
                             for n in tied_lowest]
                    print(c(f"\n  Tie for lowest: eliminating all "
                            f"would leave < 2.", Color.YELLOW + Color.BOLD))
                    print(c(f"  Tiebreak runoff ballot among: "
                            f"{', '.join(names)}", Color.YELLOW))

                    display_round_table(
                        candidate_map, vote_counts, total_active,
                        majority_needed, eliminated=tied_lowest)
                    rounds_data.append({
                        "seat_num": 1,
                        "seat_round_num": seat_round,
                        "vote_counts": dict(vote_counts),
                        "total_active": total_active,
                        "exhausted": exhausted,
                        "majority_needed": majority_needed,
                        "elected": [],
                        "eliminated": list(tied_lowest),
                        "elim_reason": ("tie for lowest — "
                                       "tiebreaker runoff ballot issued"),
                    })

                    tied_map = {n: candidate_map[n]
                                for n in tied_lowest}
                    tb_winner = run_runoff(
                        "tiebreaker", tied_map, "dei-chair",
                        "DEI Committee Chair", 1, "Tiebreaker",
                        rounds_data)
                    if tb_winner is None:
                        break

                    to_elim_tb = [n for n in tied_lowest
                                  if n != tb_winner]
                    for num in to_elim_tb:
                        active.remove(num)
                        print(c(f"  ✗ {candidate_map.get(num, num)} "
                                f"eliminated  (tiebreaker runoff)",
                                Color.RED))

                    if len(active) == 2:
                        h2h_map = {n: candidate_map[n]
                                   for n in active}
                        h2h_winner = run_runoff(
                            "head-to-head", h2h_map, "dei-chair",
                            "DEI Committee Chair", 1, "H2H",
                            rounds_data)
                        if h2h_winner:
                            winner_num = h2h_winner

                    seat_round += 1
                    continue
            else:
                to_eliminate = tied_lowest
                lowest_votes = vote_counts[tied_lowest[0]]
                lowest_pct = (lowest_votes / total_active * 100
                              if total_active > 0 else 0)
                elim_reason = (f"lowest vote-getter "
                               f"({lowest_pct:.1f}%)")

        if not winner_num and to_eliminate:
            display_round_table(candidate_map, vote_counts, total_active,
                                majority_needed, eliminated=to_eliminate)
            rounds_data.append({
                "seat_num": 1, "seat_round_num": seat_round,
                "vote_counts": dict(vote_counts),
                "total_active": total_active,
                "exhausted": exhausted,
                "majority_needed": majority_needed,
                "elected": [], "eliminated": list(to_eliminate),
                "elim_reason": elim_reason,
            })
            for num in to_eliminate:
                name = candidate_map.get(num, num)
                active.remove(num)
                print(c(f"  ✗ {name} eliminated  ({elim_reason})",
                        Color.RED))

            if len(active) == 1:
                winner_num = active[0]
                name = candidate_map.get(winner_num, winner_num)
                print(c(f"\n  Only one candidate remaining "
                        f"— {name} is elected.", Color.YELLOW))
                print(c(f"  ✓ {name} ELECTED as DEI Committee Chair",
                        Color.GREEN + Color.BOLD))

        seat_round += 1

    # Exceptions + Excel
    _print_exceptions(ballots, spoiled_surrendered,
                      election_name="DEI Committee Chair")
    if rounds_data:
        path = os.path.join(output_dir, "DEI_Chair_Results.xlsx")
        create_excel_report("DEI Committee Chair", candidate_map,
                            rounds_data, path)

    if winner_num:
        log(f"\n[FINAL] DEI Chair — Elected: "
            f"{candidate_map.get(winner_num, winner_num)}",
            also_print=False)
    return winner_num


# =============================================================================
# ELECTION C — STATE CONVENTION COMMITTEE (14 seats)   Article XII
# =============================================================================

def run_committee_election(output_dir):
    """Run the State Convention Committee election (slate vote)."""
    section_header("STATE CONVENTION COMMITTEE ELECTION — Article XII")
    log("\n[ELECTION] State Convention Committee — 14 seats (slate vote)",
        also_print=False)
    log(f"[DATE] {datetime.datetime.now().isoformat()}", also_print=False)

    seats = 14
    print(c("\n  This is a slate vote (YES / NO on the full slate).",
            Color.CYAN))
    print(c("  Nominees should already be recorded by the Committee.",
            Color.GREY))

    choice = menu("How will slate votes be loaded?", [
        "Google Sheets CSV export",
        "Paper ballot CSV  (Slate Vote column: YES or NO)",
        "Enter totals manually",
    ])

    yes_votes = 0
    no_votes  = 0
    nominee_names = []

    if choice in (1, 2):
        label = "Convention Committee — Slate Vote"
        path  = get_csv_filepath(label)
        if path:
            try:
                with open(path, newline="", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        vote = str(row.get("Slate Vote",
                                  row.get("slate_vote", ""))).strip().upper()
                        if vote in ("YES", "Y", "1"):
                            yes_votes += 1
                        elif vote in ("NO", "N", "0"):
                            no_votes += 1
            except (FileNotFoundError, OSError) as e:
                print(c(f"  Error: {e}", Color.RED))

    if choice == 3:
        try:
            yes_votes = int(input(
                c("  Enter YES votes: ", Color.CYAN)).strip())
            no_votes  = int(input(
                c("  Enter NO  votes: ", Color.CYAN)).strip())
        except ValueError:
            print(c("  Invalid numbers.", Color.RED))
            return False

    total = yes_votes + no_votes
    if total == 0:
        print(c("\n  No votes recorded.", Color.YELLOW))
        return False

    pct_yes = yes_votes / total * 100
    pct_no  = no_votes  / total * 100

    print(c(f"\n  Slate Vote Results:", Color.BOLD))
    print(c(f"    YES: {yes_votes:>5}  ({pct_yes:.1f}%)", Color.GREEN))
    print(c(f"    NO:  {no_votes:>5}  ({pct_no:.1f}%)", Color.RED))
    print(c(f"    Total: {total}", Color.WHITE))

    majority = total // 2 + 1
    approved = yes_votes >= majority

    if approved:
        print(c(f"\n  ✓ Slate APPROVED ({yes_votes} >= {majority} needed)",
                Color.GREEN + Color.BOLD))
        log(f"[RESULT] Slate APPROVED: YES {yes_votes}, NO {no_votes}",
            also_print=False)
    else:
        print(c(f"\n  ✗ Slate FAILED ({yes_votes} < {majority} needed)",
                Color.RED + Color.BOLD))
        log(f"[RESULT] Slate FAILED: YES {yes_votes}, NO {no_votes}",
            also_print=False)

    return approved


# =============================================================================
# EXCEPTIONS REPORT
# =============================================================================

def _print_exceptions(ballots, spoiled_surrendered, election_name=""):
    exhausted     = sum(1 for b in ballots if not b.is_spoiled
                        and b.get_first_active_choice([]) is None)
    n_total       = len(spoiled_surrendered)
    n_surrendered = sum(1 for b in spoiled_surrendered
                        if "Non-Issued" not in (b.spoil_reason or ""))
    n_nonissued   = sum(1 for b in spoiled_surrendered
                        if "Non-Issued" in (b.spoil_reason or ""))

    if n_total == 0 and exhausted == 0:
        return

    print(c(f"\n  ── Exceptions — {election_name} ──", Color.YELLOW))
    log(f"\n[EXCEPTIONS] {election_name}", also_print=False)
    if n_surrendered:
        log(f"  Surrendered Creds : {n_surrendered}", also_print=False)
        print(c(f"    Surrendered Credentials : {n_surrendered}", Color.YELLOW))
    if n_nonissued:
        log(f"  Non-Issued Creds  : {n_nonissued}", also_print=False)
        print(c(f"    Non-Issued Credentials  : {n_nonissued}", Color.RED))
    if exhausted:
        log(f"  Exhausted         : {exhausted}", also_print=False)
        print(c(f"    Exhausted Ballots       : {exhausted}", Color.GREY))
    for b in spoiled_surrendered:
        tag = "NON-ISSUED" if "Non-Issued" in (b.spoil_reason or "") else "SURRENDERED"
        log(f"  [{tag}] Delegate {b.delegate_number} — "
            f"{b.spoil_reason}", also_print=False)


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def create_excel_report(election_name, candidate_map, rounds_data,
                        output_path):
    """Write an Excel workbook with Summary + per-round sheets."""
    if not EXCEL_AVAILABLE:
        print(c("\n  openpyxl not installed — skipping Excel export.",
                Color.YELLOW))
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_fill   = PatternFill("solid", fgColor="1F3864")
    elect_fill = PatternFill("solid", fgColor="C6EFCE")
    elim_fill  = PatternFill("solid", fgColor="FFCCCC")
    hdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    bold_font  = Font(name="Calibri", bold=True, size=11)
    norm_font  = Font(name="Calibri", size=11)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    def sc(cell, fill=None, font=None, align="left"):
        if fill:
            cell.fill = fill
        cell.font      = font or norm_font
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = thin

    # ── Per-round sheets ────────────────────────────────────────────────
    for rd in rounds_data:
        seat_n  = rd.get("seat_num", "")
        seat_rn = rd.get("seat_round_num", 1)

        if seat_n:
            if isinstance(seat_rn, int):
                tab_title  = f"Seat {seat_n} - Round {seat_rn}"
                full_title = (f"Seat {seat_n} Ballot — "
                              f"Distribution Round {seat_rn}")
            else:
                tab_title  = f"Seat {seat_n} - {seat_rn}"
                full_title = f"Seat {seat_n} Ballot — {seat_rn} Runoff"
        else:
            tab_title  = f"Round {seat_rn}"
            full_title = f"Distribution Round {seat_rn}"

        ws = wb.create_sheet(title=tab_title[:31])

        ws.merge_cells("A1:F1")
        ws["A1"] = f"{election_name} — {full_title}"
        ws["A1"].font = Font(name="Calibri", bold=True, size=13,
                             color="1F3864")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:F2")
        ws["A2"] = (f"Active: {rd['total_active']}   |   "
                    f"Majority: {rd['majority_needed']}   |   "
                    f"15% threshold: "
                    f"{rd['total_active'] * 0.15:.1f}   |   "
                    f"Exhausted: {rd.get('exhausted', 0)}")
        ws["A2"].font = Font(name="Calibri", italic=True, size=10,
                             color="595959")
        ws["A2"].alignment = Alignment(horizontal="center")

        elim_rule = rd.get("elim_reason", "")
        if not elim_rule:
            srv = rd.get("seat_round_num", 1)
            if srv == 1:
                elim_rule = "15% elimination threshold"
            elif isinstance(srv, str):
                elim_rule = f"{srv} runoff ballot"
            else:
                elim_rule = "Lowest vote-getter eliminated"
        ws.merge_cells("A3:F3")
        ws["A3"] = f"Elimination rule: {elim_rule}"
        ws["A3"].font = Font(name="Calibri", italic=True, size=10,
                             color="7F5A00")
        ws["A3"].alignment = Alignment(horizontal="center")

        headers = ["#", "Candidate", "Votes", "% of Active",
                   "Majority?", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            sc(cell, fill=hdr_fill, font=hdr_font, align="center")

        sorted_cands = sorted(rd["vote_counts"].items(),
                              key=lambda x: -x[1])
        for row_i, (num, votes) in enumerate(sorted_cands, 6):
            name = candidate_map.get(num, f"Candidate {num}")
            pct = (votes / rd["total_active"] * 100
                   if rd["total_active"] > 0 else 0)
            maj = ("YES ✓" if votes >= rd["majority_needed"]
                   else "No")

            if num in rd["elected"]:
                status = "ELECTED"
                fill, font = elect_fill, bold_font
            elif num in rd["eliminated"]:
                status = f"ELIMINATED — {rd.get('elim_reason','')}"
                fill, font = elim_fill, norm_font
            else:
                status = "Continuing"
                fill, font = None, norm_font

            vals = [num, name, votes, f"{pct:.1f}%", maj, status]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                sc(cell, fill=fill, font=font,
                   align="center" if col != 2 else "left")

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 30
        for ch in ["C", "D", "E", "F"]:
            ws.column_dimensions[ch].width = 16

    # ── Summary sheet ────────────────────────────────────────────────────
    ws_s = wb.create_sheet(title="Summary", index=0)
    ws_s.merge_cells("A1:E1")
    ws_s["A1"] = f"{election_name} — Final Results"
    ws_s["A1"].font = Font(name="Calibri", bold=True, size=14,
                           color="1F3864")
    ws_s["A1"].alignment = Alignment(horizontal="center")

    ws_s.merge_cells("A2:E2")
    ws_s["A2"] = (f"Generated: "
                  f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
                  f"   |   {DISTRICT}")
    ws_s["A2"].font = Font(name="Calibri", italic=True, size=10,
                           color="595959")
    ws_s["A2"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(
            ["#", "Candidate", "Seat", "Elected on Round", "Votes"], 1):
        cell = ws_s.cell(row=4, column=col, value=h)
        sc(cell, fill=hdr_fill, font=hdr_font, align="center")

    elected_summary = []
    for rd in rounds_data:
        for num in rd["elected"]:
            votes   = rd["vote_counts"].get(num, 0)
            seat_n  = rd.get("seat_num", "")
            seat_rn = rd.get("seat_round_num", "")
            elected_summary.append((num, seat_n, seat_rn, votes))

    for row_i, (num, seat, seat_rnd, votes) in enumerate(
            elected_summary, 5):
        name = candidate_map.get(num, f"Candidate {num}")
        vals = [row_i - 4, name,
                f"Seat {seat}" if seat else "", seat_rnd, votes]
        for col, val in enumerate(vals, 1):
            cell = ws_s.cell(row=row_i, column=col, value=val)
            sc(cell, fill=elect_fill,
               align="center" if col != 2 else "left")

    ws_s.column_dimensions["A"].width = 5
    ws_s.column_dimensions["B"].width = 30
    for ch in ["C", "D", "E"]:
        ws_s.column_dimensions[ch].width = 20

    wb.save(output_path)
    print(c(f"\n  Excel report saved → {output_path}", Color.GREEN))
    log(f"[EXCEL] Saved: {output_path}", also_print=False)


# =============================================================================
# OUTPUT DIRECTORY
# =============================================================================

def get_output_dir(script_dir):
    """Prompt operator to select where results are saved."""
    default_local = os.path.join(script_dir, "election_results")

    drive_paths = []
    for candidate in [
        os.path.expanduser("~/Google Drive/My Drive"),
        os.path.expanduser("~/Google Drive"),
        "G:\\My Drive",
        "G:\\",
    ]:
        if os.path.isdir(candidate):
            drive_paths.append(candidate)

    print(c(f"\n  Where should election result files be saved?",
            Color.BOLD))
    options = [f"Local folder: {default_local}"]
    if drive_paths:
        options.append(f"Google Drive (detected: {drive_paths[0]})")
    options.append("Enter a custom folder path")

    choice = menu("Select output location:", options)

    if choice == 1:
        output_dir = default_local
    elif choice == 2 and drive_paths:
        drive_root = drive_paths[0]
        convention_folder = os.path.join(
            drive_root, "3rd CD Convention", "election_results")
        print(c(f"\n  Will save to: {convention_folder}", Color.CYAN))
        if not confirm("  Use this Drive folder?"):
            custom = input("  Enter full Drive folder path: ").strip()
            custom = custom.strip('"').strip("'")
            if custom:
                convention_folder = custom
        output_dir = convention_folder
    else:
        custom = input("  Enter full folder path: ").strip()
        custom = custom.strip('"').strip("'")
        output_dir = custom if custom else default_local

    os.makedirs(output_dir, exist_ok=True)
    print(c(f"\n  Results will be saved to:", Color.BOLD))
    print(c(f"    {output_dir}", Color.GREEN))
    return output_dir


# =============================================================================
# MAIN
# =============================================================================

def main():
    banner()

    log(f"SESSION STARTED: {datetime.datetime.now().isoformat()}",
        also_print=False)
    log(f"CONVENTION: {DISTRICT} — {CONVENTION_DATE}", also_print=False)
    log(f"TABULATOR VERSION: {VERSION}", also_print=False)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = get_output_dir(script_dir)
    log(f"OUTPUT DIR: {output_dir}", also_print=False)

    all_results = {}

    while True:
        section_header("MAIN MENU")
        choice = menu("Select an election to run:", [
            "A.  State Central Committee — Women           (4 seats, IRV)",
            "B.  State Central Committee — Men/Non-Binary  (4 seats, IRV)",
            "C.  DEI Committee Chair                        (1 seat, IRV)",
            "D.  State Convention Committee                 (14 seats, slate)",
            "Save audit log and exit",
        ])

        if choice == 1:
            result = load_ballots_for_election("SCC — Women", script_dir)
            if result[0] is not None:
                ballots, cmap, spoiled = result
                elected, _, _ = run_scc_election(
                    "Women", 4, ballots, cmap, spoiled, output_dir)
                all_results["SCC_Women"] = [cmap.get(n, n)
                                            for n in elected]

        elif choice == 2:
            result = load_ballots_for_election(
                "SCC — Men/Non-Binary", script_dir)
            if result[0] is not None:
                ballots, cmap, spoiled = result
                elected, _, _ = run_scc_election(
                    "Men_NonBinary", 4, ballots, cmap, spoiled,
                    output_dir)
                all_results["SCC_MNB"] = [cmap.get(n, n)
                                          for n in elected]

        elif choice == 3:
            result = load_ballots_for_election(
                "DEI Committee Chair", script_dir)
            if result[0] is not None:
                ballots, cmap, spoiled = result
                winner = run_dei_election(
                    ballots, cmap, spoiled, output_dir)
                all_results["DEI_Chair"] = (
                    cmap.get(winner, winner) if winner else None)

        elif choice == 4:
            slate_approved = run_committee_election(output_dir)
            all_results["Convention_Committee_Slate"] = (
                "Approved" if slate_approved else "Failed")

        elif choice == 5:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            log_path = os.path.join(output_dir,
                                    f"audit_log_{timestamp}.txt")
            save_audit_log(log_path)
            print(c("\n  ══ SESSION COMPLETE ══",
                    Color.BOLD + Color.CYAN))
            print(c("\n  Elections completed this session:",
                    Color.WHITE))
            for k, v in all_results.items():
                print(c(f"    {k}: {v}", Color.GREEN))
            print(c(f"\n  All files saved to: {output_dir}",
                    Color.CYAN))
            break


if __name__ == "__main__":
    main()
