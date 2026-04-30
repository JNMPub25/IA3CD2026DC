#!/usr/bin/env python3
"""
=============================================================================
3rd Congressional District Democratic Convention — May 2, 2026
ELECTION TABULATION TOOL  v2.0  —  Ranked-Choice IRV Tabulator
=============================================================================
Handles three elections:
  A. State Central Committee (SCC)   — Sections V.K / VI    (IRV, 4 seats each)
  B. DEI Committee Chair             — Sections V.K / IX    (IRV, 1 seat)
  C. State Convention Committee      — Sections V.K / XII   (Slate vote, 14 seats)

Ballot input:
  • Google Sheets CSV export   (primary — electronic ballots)
  • Paper ballot CSV           (optional — merged by delegate number)

Other inputs:
  • surrendered_delegates.json  (produced by district3_setup.py)

Outputs:
  • Round-by-round results on screen
  • Excel workbook  (Summary + one sheet per round per election)
  • Plain-text audit log  (for the official record)
  • Exceptions report  (Spoiled / Exhausted / Surrendered Credentials)

Requirements:  Python 3.8+,  openpyxl  (pip install openpyxl)
=============================================================================

SCC SEAT-RESTART RULE (confirmed with IDP contact 2026-04-28):
  After each SCC seat is won, ONLY the winner is removed from the pool.
  All other candidates — including those eliminated during that seat's
  rounds — return to the pool for the next seat.  Example:
    Seat 1: A wins, C & D eliminated → Seat 2 pool: B, C, D, E, F, ...
    Seat 2: B wins, E eliminated     → Seat 3 pool: C, D, E, F, ...
  Each seat starts a fresh IRV cycle with the full non-elected pool.
=============================================================================
"""

import csv
import json
import os
import re
import sys
import datetime
from collections import defaultdict
from pathlib import Path

# ── Try to import openpyxl ──────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# =============================================================================
# CONSTANTS
# =============================================================================

VERSION           = "2.0.0"
CONVENTION_DATE   = "May 2, 2026"
DISTRICT          = "3rd Congressional District Democratic Convention"

ELIMINATION_THRESHOLD   = 0.15   # 15%  — candidates below this are eliminated
SCC_FIRST_BALLOT_MAX    = 0.50   # 50%  — max fraction of seats filled in Round 1
MAJORITY_THRESHOLD      = 0.50   # must receive MORE THAN 50% to be elected

# Column header patterns in the Google Sheets CSV export
# e.g. "Candidate A Rank", "Candidate B Rank"
RANK_HEADER_RE = re.compile(r'candidate\s+([A-Za-z])\s+rank', re.IGNORECASE)

# =============================================================================
# COLOR / DISPLAY HELPERS
# =============================================================================

class Color:
    RESET   = "\033[0m"
    BOLD    = "\033[1m"
    GREEN   = "\033[92m"
    YELLOW  = "\033[93m"
    RED     = "\033[91m"
    CYAN    = "\033[96m"
    MAGENTA = "\033[95m"
    WHITE   = "\033[97m"
    GREY    = "\033[90m"

def c(text, color):
    return f"{color}{text}{Color.RESET}"

def banner():
    w = 70
    print()
    print(c("=" * w, Color.CYAN))
    print(c(f"  {DISTRICT}".center(w), Color.BOLD + Color.WHITE))
    print(c(f"  Election Tabulation Tool  v{VERSION}".center(w), Color.CYAN))
    print(c(f"  {CONVENTION_DATE}".center(w), Color.GREY))
    print(c("=" * w, Color.CYAN))
    print()

def section_header(title):
    print()
    print(c("─" * 62, Color.CYAN))
    print(c(f"  {title}", Color.BOLD + Color.WHITE))
    print(c("─" * 62, Color.CYAN))
    log(f"\n{'─'*62}\n  {title}\n{'─'*62}", also_print=False)

def seat_ballot_header(seat_num, election_label):
    """Displayed once when a seat's tabulation begins."""
    msg = f"  SEAT {seat_num} BALLOT — {election_label}"
    print()
    print(c("╔" + "═" * 60 + "╗", Color.MAGENTA + Color.BOLD))
    print(c(f"║{msg:<60}║", Color.MAGENTA + Color.BOLD))
    print(c("╚" + "═" * 60 + "╝", Color.MAGENTA + Color.BOLD))
    log(f"\n{'═'*62}\n[SEAT {seat_num} BALLOT] {election_label}\n{'═'*62}", also_print=False)

def distribution_round_header(seat_round_num, seat_num, election_label):
    """Displayed for each redistribution round within a seat's ballot."""
    msg = f"  Seat {seat_num} Ballot — Distribution Round {seat_round_num}"
    print()
    print(c("┌" + "─" * 60 + "┐", Color.YELLOW))
    print(c(f"│{msg:<60}│", Color.YELLOW))
    print(c("└" + "─" * 60 + "┘", Color.YELLOW))
    log(f"\n[SEAT {seat_num} — Distribution Round {seat_round_num}] {election_label}", also_print=False)

def display_round_table(candidate_map, vote_counts, total_active, majority_needed,
                        elected=None, eliminated=None):
    """
    Print a formatted table of vote counts for one IRV round.
    candidate_map : {cand_num_str : name_str}
    vote_counts   : {cand_num_str : int}
    """
    elected    = elected    or []
    eliminated = eliminated or []

    threshold_votes = total_active * ELIMINATION_THRESHOLD

    sorted_cands = sorted(vote_counts.items(), key=lambda x: -x[1])

    col_w = [32, 8, 10, 16]
    header = (f"  {'Candidate':<{col_w[0]}} {'Votes':>{col_w[1]}} "
              f"{'  %':>{col_w[2]}} {'Status':<{col_w[3]}}")
    divider = "  " + "─" * (sum(col_w) + 3 * 2)

    print(c(header, Color.BOLD))
    print(c(divider, Color.GREY))
    log(header, also_print=False)
    log(divider, also_print=False)

    for num, votes in sorted_cands:
        name = candidate_map.get(num, f"Candidate {num}")
        pct  = (votes / total_active * 100) if total_active > 0 else 0

        if num in elected:
            status = "✓ ELECTED"
            color  = Color.GREEN
        elif num in eliminated:
            status = "✗ ELIMINATED"
            color  = Color.RED
        else:
            status = ""
            color  = Color.WHITE

        row = (f"  {name:<{col_w[0]}} {votes:>{col_w[1]}} "
               f"{pct:>{col_w[2]-1}.1f}%  {status:<{col_w[3]}}")
        print(c(row, color))
        log(row, also_print=False)

    summary = (f"\n  Active ballots   : {total_active}"
               f"\n  Majority needed  : {majority_needed}  (> 50% of active ballots)"
               f"\n  15% threshold    : {threshold_votes:.1f} votes")
    print(c(divider, Color.GREY))
    print(c(summary, Color.GREY))
    log(divider + summary, also_print=False)
    print()

def prompt(msg, default=None):
    suffix = f" [{default}]" if default is not None else ""
    return input(f"\n  {msg}{suffix}: ").strip() or (str(default) if default is not None else "")

def confirm(msg):
    ans = input(f"\n  {msg} (y/n): ").strip().lower()
    return ans in ("y", "yes")

def menu(title, options):
    print()
    print(c(f"  {title}", Color.BOLD))
    for i, opt in enumerate(options, 1):
        print(f"    {c(str(i), Color.CYAN)}.  {opt}")
    while True:
        try:
            choice = int(input("\n  Enter number: ").strip())
            if 1 <= choice <= len(options):
                return choice
        except ValueError:
            pass
        print(c("  Please enter a valid number.", Color.RED))

# =============================================================================
# AUDIT LOG
# =============================================================================

_audit_lines = []

def log(text, also_print=True):
    _audit_lines.append(text)
    if also_print:
        print(text)

def save_audit_log(filepath):
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(_audit_lines))
    print(c(f"\n  Audit log saved → {filepath}", Color.GREEN))

# =============================================================================
# BALLOT DATA CLASS
# =============================================================================

class Ballot:
    """Represents one delegate's ranked-choice ballot."""
    __slots__ = ("delegate_number", "timestamp", "rankings",
                 "source", "is_spoiled", "spoil_reason")

    def __init__(self, delegate_number, timestamp, rankings, source="electronic"):
        self.delegate_number = str(delegate_number).strip()
        self.timestamp       = timestamp
        # rankings: {candidate_num_str: rank_int}
        # rank_int == 0 means "not ranked" (treat as unranked)
        self.rankings  = rankings
        self.source    = source        # "electronic" or "paper"
        self.is_spoiled  = False
        self.spoil_reason = ""

    def get_first_active_choice(self, active_keys):
        """
        Return the candidate letter of this ballot's highest-ranked (lowest rank int)
        candidate who is still active.  Returns None if ballot is exhausted.
        active_keys : iterable of candidate letter strings
        """
        best_num  = None
        best_rank = None
        for num in active_keys:
            rank = self.rankings.get(num)
            if rank and isinstance(rank, int) and rank > 0:
                if best_rank is None or rank < best_rank:
                    best_rank = rank
                    best_num  = num
        return best_num

# =============================================================================
# CSV LOADING — ELECTRONIC BALLOTS (Google Sheets export)
# =============================================================================
#
# Expected column headers (set by apps_script.js writeHeader):
#   Timestamp | Delegate Number | Election | Is Test | Is Auto-Submit |
#   Candidate A Rank | Candidate B Rank | Candidate C Rank | ...
#
# Each "Candidate A Rank" cell contains an integer (the rank the delegate
# assigned to that candidate) or is blank (candidate not ranked).

def load_electronic_ballots(filepath):
    """
    Load ranked-choice ballots from a Google Sheets CSV export.
    Returns (ballots: list[Ballot], candidate_keys: list[str], warnings: list[str])
    where candidate_keys are the letters found in the column headers.
    """
    ballots        = []
    warnings       = []
    candidate_keys = []

    try:
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader     = csv.DictReader(f)
            fieldnames = [fn.strip() for fn in (reader.fieldnames or [])]

            # Discover candidate letter columns
            rank_cols = {}   # {candidate_letter_str: exact_header}
            for fn in fieldnames:
                m = RANK_HEADER_RE.match(fn.strip())
                if m:
                    rank_cols[m.group(1).upper()] = fn

            if not rank_cols:
                warnings.append("No 'Candidate A Rank' columns found in this CSV. "
                                 "Are you using the Google Sheets export?")
                return [], [], warnings

            candidate_keys = sorted(rank_cols.keys())

            for row_num, row in enumerate(reader, start=2):
                delegate_num = str(row.get("Delegate Number", row.get("delegate_number", ""))).strip()
                timestamp    = str(row.get("Timestamp", "")).strip()
                is_test      = str(row.get("Is Test", "NO")).strip().upper()

                if is_test == "YES":
                    continue   # skip test submissions

                # Parse rankings
                rankings = {}
                for cand_num, col in rank_cols.items():
                    raw = str(row.get(col, "")).strip()
                    if raw == "":
                        rankings[cand_num] = 0   # not ranked
                    else:
                        try:
                            rankings[cand_num] = int(raw)
                        except ValueError:
                            rankings[cand_num] = 0
                            warnings.append(f"Row {row_num}: non-integer rank '{raw}' "
                                            f"for Candidate {cand_num} (delegate {delegate_num}) — ignored")

                if not delegate_num:
                    warnings.append(f"Row {row_num}: missing delegate number — ballot skipped")
                    continue

                ballots.append(Ballot(delegate_num, timestamp, rankings, source="electronic"))

    except FileNotFoundError:
        warnings.append(f"File not found: {filepath}")

    return ballots, candidate_keys, warnings

# =============================================================================
# CSV LOADING — PAPER BALLOTS
# =============================================================================
#
# Paper ballot CSV uses the same format as the Google Sheets export:
#   Ballot_Number, Delegate_Number, Candidate_A_Rank, Candidate_B_Rank, ...
#
# "Candidate_A_Rank" columns contain the rank (1, 2, 3...) or blank.
# Column names are flexible — any column matching "Candidate A Rank" (case-
# insensitive) is treated as a rank column.

def load_paper_ballots(filepath):
    """
    Load paper ballots from CSV.
    Returns (ballots: list[Ballot], candidate_keys: list[str], warnings: list[str])
    """
    # Re-uses the same loader as electronic; paper format mirrors Google Sheets
    ballots, candidate_keys, warnings = load_electronic_ballots(filepath)
    for b in ballots:
        b.source = "paper"
    return ballots, candidate_keys, warnings

# =============================================================================
# BALLOT MERGE
# =============================================================================

def _count_rankings(ballot):
    """Count how many candidates a ballot actually ranked (rank > 0)."""
    return sum(1 for r in ballot.rankings.values() if isinstance(r, int) and r > 0)


def merge_ballots(electronic, paper):
    """
    Merge electronic and paper ballots.

    Priority: PAPER over electronic, UNLESS the paper ballot has fewer
    rankings than the electronic one.  In that case, whichever ballot
    holds the most rankings is accepted.

    Returns merged list[Ballot].
    """
    by_delegate = {}

    # Index electronic ballots first
    electronic_by_dn = {}
    for b in electronic:
        # If same delegate submitted multiple electronic ballots, keep last
        electronic_by_dn[b.delegate_number] = b

    # Index paper ballots
    paper_by_dn = {}
    for b in paper:
        paper_by_dn[b.delegate_number] = b

    # All unique delegate numbers
    all_delegates = set(electronic_by_dn.keys()) | set(paper_by_dn.keys())

    paper_used = 0
    electronic_used = 0
    conflicts = 0

    for dn in all_delegates:
        e_ballot = electronic_by_dn.get(dn)
        p_ballot = paper_by_dn.get(dn)

        if p_ballot and not e_ballot:
            by_delegate[dn] = p_ballot
            paper_used += 1
        elif e_ballot and not p_ballot:
            by_delegate[dn] = e_ballot
            electronic_used += 1
        else:
            # Both exist — paper wins UNLESS it has fewer rankings
            conflicts += 1
            p_ranks = _count_rankings(p_ballot)
            e_ranks = _count_rankings(e_ballot)
            if p_ranks >= e_ranks:
                by_delegate[dn] = p_ballot
                paper_used += 1
                log(f"  [MERGE] Delegate {dn}: paper ballot used "
                    f"({p_ranks} rankings vs {e_ranks} electronic).",
                    also_print=False)
            else:
                by_delegate[dn] = e_ballot
                electronic_used += 1
                log(f"  [MERGE] Delegate {dn}: electronic ballot used "
                    f"— paper had fewer rankings ({p_ranks} vs {e_ranks}).",
                    also_print=False)

    merged = list(by_delegate.values())
    log(f"\n[MERGE] Electronic submitted: {len(electronic)}  "
        f"Paper submitted: {len(paper)}  "
        f"Conflicts: {conflicts}  Total merged: {len(merged)}",
        also_print=False)
    return merged

# =============================================================================
# SURRENDERED CREDENTIALS CHECK
# =============================================================================

def _parse_timestamp(ts_str):
    """
    Parse a timestamp string from either:
      - ISO format:         "2026-05-02T09:30:00" or "2026-05-02 09:30:00"
      - toLocaleString():   "5/2/2026, 9:35:00 AM"  (US locale from Google Apps Script)
    Returns a datetime or None if unparseable.
    """
    if not ts_str or not isinstance(ts_str, str):
        return None
    ts_str = ts_str.strip()

    # Try ISO first
    try:
        return datetime.datetime.fromisoformat(ts_str)
    except ValueError:
        pass

    # Try toLocaleString() format: "M/D/YYYY, H:MM:SS AM/PM"
    for fmt in ("%m/%d/%Y, %I:%M:%S %p",   # 5/2/2026, 9:35:00 AM
                "%m/%d/%Y %I:%M:%S %p",     # 5/2/2026 9:35:00 AM  (no comma)
                "%m/%d/%Y, %H:%M:%S",       # 5/2/2026, 09:35:00   (24-hr)
                "%m/%d/%Y %H:%M:%S"):       # 5/2/2026 09:35:00
        try:
            return datetime.datetime.strptime(ts_str, fmt)
        except ValueError:
            continue

    return None


def check_surrendered(ballots, surrendered_path):
    """
    Load surrendered_delegates.json and flag any ballot from a delegate who
    surrendered their credentials as "Spoiled — Surrendered Credentials".

    Any ballot submitted by a delegate in the surrendered list is spoiled
    regardless of timing — if they surrendered, their vote shouldn't count.

    surrendered_delegates.json format (created by district3_setup.py):
      [ { "delegate_number": "101", "name": "Jane Doe",
          "surrender_time": "2026-05-02 09:30:00" }, ... ]

    Returns (valid_ballots, spoiled_surrendered) lists.
    """
    surrendered_set = set()    # delegate numbers who surrendered
    surrendered_info = {}      # {delegate_number: {"name": ..., "time": ...}}

    if not os.path.isfile(surrendered_path):
        log("  [SURRENDERED] No surrendered_delegates.json found — skipping check.",
            also_print=False)
        return ballots, []

    try:
        with open(surrendered_path, encoding="utf-8") as f:
            records = json.load(f)
        for rec in records:
            dn = str(rec.get("delegate_number", "")).strip()
            if dn:
                surrendered_set.add(dn)
                surrendered_info[dn] = {
                    "name": rec.get("name", "Unknown"),
                    "time": rec.get("surrender_time", ""),
                }
    except (json.JSONDecodeError, OSError) as e:
        log(f"  [WARNING] Could not read surrendered_delegates.json: {e}",
            also_print=False)
        return ballots, []

    if not surrendered_set:
        log("  [SURRENDERED] File loaded but no delegate entries found.",
            also_print=False)
        return ballots, []

    log(f"  [SURRENDERED] Loaded {len(surrendered_set)} surrendered delegate(s).",
        also_print=False)

    valid   = []
    spoiled = []

    for b in ballots:
        if b.delegate_number in surrendered_set:
            info = surrendered_info[b.delegate_number]
            surrender_ts = info["time"]
            ballot_dt = _parse_timestamp(b.timestamp)
            surrender_dt = _parse_timestamp(surrender_ts)

            time_detail = ""
            if ballot_dt and surrender_dt:
                time_detail = (f" (surrendered {surrender_dt.strftime('%H:%M')}, "
                               f"voted {ballot_dt.strftime('%H:%M')})")
            elif surrender_dt:
                time_detail = f" (surrendered {surrender_dt.strftime('%H:%M')})"

            b.is_spoiled   = True
            b.spoil_reason = f"Surrendered Credentials{time_detail}"
            spoiled.append(b)
            log(f"  [SPOILED] Delegate {b.delegate_number} ({info['name']}) "
                f"— {b.spoil_reason}", also_print=False)
        else:
            valid.append(b)

    return valid, spoiled

# =============================================================================
# CANDIDATE SETUP
# =============================================================================

def enter_candidates(election_name, found_keys=None):
    """
    Prompt operator to enter candidate names by letter.
    found_keys: list of candidate letter strings found in the CSV headers.
    Returns dict {letter_str: name_str}.
    """
    print()
    print(c(f"  Enter candidate names for: {election_name}", Color.BOLD))
    if found_keys:
        print(c(f"  (CSV has {len(found_keys)} candidate columns: "
                f"Candidates {', '.join(found_keys)})", Color.GREY))
        print(c("  Enter each candidate's name in the same lettered order.", Color.GREY))

    candidate_map = {}

    if found_keys:
        for key in found_keys:
            while True:
                name = input(f"    Name for Candidate {key}: ").strip()
                if name:
                    candidate_map[key] = name
                    break
                print(c("    Name cannot be blank.", Color.RED))
    else:
        # Manual entry (no CSV loaded yet)
        print(c("  Press Enter on a blank line when done.", Color.GREY))
        idx = 0
        while True:
            letter = chr(65 + idx)   # A, B, C, …
            name = input(f"    Candidate {letter} name (or Enter to finish): ").strip()
            if not name:
                if idx < 2:
                    print(c("  At least 2 candidates required.", Color.RED))
                    continue
                break
            candidate_map[letter] = name
            idx += 1

    log(f"\n[CANDIDATES — {election_name}]", also_print=False)
    for key, name in candidate_map.items():
        log(f"  {key}: {name}", also_print=False)

    print(c(f"\n  Confirmed {len(candidate_map)} candidates:", Color.CYAN))
    for key, name in candidate_map.items():
        print(f"    {key}. {name}")

    return candidate_map

def get_csv_filepath(label):
    """Prompt for a CSV path and validate it exists. Returns path or None."""
    print(c(f"\n  Enter path to CSV file for {label}:", Color.BOLD))
    print(c("  (You can drag the file into this window)", Color.GREY))
    while True:
        path = input("  Path: ").strip().strip('"').strip("'")
        if os.path.isfile(path):
            return path
        print(c(f"  File not found: {path}", Color.RED))
        if not confirm("  Try again?"):
            return None

def load_ballots_for_election(election_name, data_dir):
    """
    Interactive ballot loading:
    1. Ask for Google Sheets CSV (required unless paper-only).
    2. Optionally ask for paper ballot CSV.
    3. Merge; check surrendered credentials.
    Returns (ballots, candidate_map) or (None, None) on failure.
    """
    section_header(f"BALLOT LOADING — {election_name}")

    electronic_ballots = []
    paper_ballots      = []
    candidate_keys     = []
    warnings           = []

    # ── Electronic (Google Sheets) ───────────────────────────────────────────
    choice = menu("How will ballots be loaded?", [
        "Google Sheets CSV export  (recommended — electronic ballots)",
        "Paper ballot CSV only  (no electronic ballots)",
        "Both  (merge electronic + paper)",
    ])

    if choice in (1, 3):
        path = get_csv_filepath(f"{election_name} — Google Sheets export")
        if path:
            electronic_ballots, candidate_keys, warnings = load_electronic_ballots(path)
            for w in warnings:
                print(c(f"  ⚠  {w}", Color.YELLOW))
            print(c(f"\n  Loaded {len(electronic_ballots)} electronic ballot(s).", Color.GREEN))
            log(f"[LOAD] Electronic CSV: {path}  ({len(electronic_ballots)} ballots)", also_print=False)

    if choice in (2, 3):
        path = get_csv_filepath(f"{election_name} — paper ballots CSV")
        if path:
            paper_ballots, paper_keys, warnings = load_paper_ballots(path)
            for w in warnings:
                print(c(f"  ⚠  {w}", Color.YELLOW))
            print(c(f"\n  Loaded {len(paper_ballots)} paper ballot(s).", Color.GREEN))
            log(f"[LOAD] Paper CSV: {path}  ({len(paper_ballots)} ballots)", also_print=False)
            if not candidate_keys:
                candidate_keys = paper_keys

    if not electronic_ballots and not paper_ballots:
        print(c("\n  No ballots loaded — cannot tabulate.", Color.RED))
        return None, None

    # ── Merge ────────────────────────────────────────────────────────────────
    all_ballots = merge_ballots(electronic_ballots, paper_ballots)

    # ── Surrendered credentials ──────────────────────────────────────────────
    surrendered_path = os.path.join(data_dir, "surrendered_delegates.json")
    all_ballots, spoiled_surrendered = check_surrendered(all_ballots, surrendered_path)
    if spoiled_surrendered:
        print(c(f"\n  ⚠  {len(spoiled_surrendered)} ballot(s) flagged: Surrendered Credentials.", Color.YELLOW))
        for b in spoiled_surrendered:
            print(c(f"     Delegate {b.delegate_number} — {b.spoil_reason}", Color.YELLOW))

    # ── Candidate names ──────────────────────────────────────────────────────
    candidate_map = enter_candidates(election_name, candidate_keys)

    return all_ballots, candidate_map, spoiled_surrendered

# =============================================================================
# IRV CORE — VOTE COUNTING
# =============================================================================

def count_round_votes(ballots, active_nums):
    """
    Count first-active-choice votes for one IRV round.
    Returns:
      vote_counts   : {cand_num_str: int}   (only active candidates)
      total_active  : int  (ballots with an active choice — used for majority/threshold)
      exhausted     : int  (ballots with no active choice remaining)
    """
    vote_counts  = {num: 0 for num in active_nums}
    exhausted    = 0

    for b in ballots:
        if b.is_spoiled:
            continue
        choice = b.get_first_active_choice(active_nums)
        if choice is None:
            exhausted += 1
        else:
            vote_counts[choice] = vote_counts.get(choice, 0) + 1

    total_active = sum(vote_counts.values())
    return vote_counts, total_active, exhausted

def compute_threshold_elimination(vote_counts, threshold=ELIMINATION_THRESHOLD):
    """
    STEP 1 ONLY — Apply the 15% threshold rule.

    Eliminates candidates with < 15% of total active votes, subject to the
    safeguard that no more than 49% of remaining candidates may be eliminated.

    Rules (Convention Rules Article V.K):
      1. Eliminate candidates with < 15% of total active votes.
      2. But if that would eliminate ≥ 50% of remaining candidates,
         lower the cut until fewer than 50% would be eliminated.

    This function does NOT drop the lowest vote-getter.  After calling this,
    the caller should redistribute votes and check for a 50%+1 winner before
    proceeding to drop_lowest_candidate() as a separate step.

    Returns: (to_eliminate: list[str], actual_threshold_pct: float)
    """
    total = sum(vote_counts.values())
    if total == 0:
        return [], 0.0

    n = len(vote_counts)
    sorted_cands = sorted(vote_counts.items(), key=lambda x: x[1])

    # Step 1: Apply threshold
    cut_votes       = total * threshold
    below_threshold = [num for num, v in vote_counts.items() if v < cut_votes]
    actual_pct      = threshold * 100

    # Step 2: Safeguard — cannot eliminate ≥ 50% of candidates
    max_eliminate = max(1, (n - 1) // 2)

    if len(below_threshold) > max_eliminate:
        below_threshold = [num for num, v in sorted_cands[:max_eliminate]]
        if below_threshold:
            cut_votes  = vote_counts[below_threshold[-1]]
            actual_pct = cut_votes / total * 100

    return below_threshold, actual_pct


def drop_lowest_candidate(vote_counts):
    """
    STEP 3 — Drop the single lowest vote-getter.

    Called only when the 15% threshold step did not produce a winner.
    Returns the candidate number string of the lowest vote-getter.
    If there is a tie for lowest, returns ALL tied candidates (the caller
    should invoke a tiebreak runoff if eliminating all would leave < 2).
    """
    if not vote_counts:
        return []
    sorted_cands = sorted(vote_counts.items(), key=lambda x: x[1])
    lowest_votes = sorted_cands[0][1]
    tied_lowest  = [num for num, v in vote_counts.items() if v == lowest_votes]
    return tied_lowest


def compute_elimination_cut(vote_counts, threshold=ELIMINATION_THRESHOLD):
    """
    LEGACY WRAPPER — Bundled threshold + lowest-drop in one call.

    Retained for reference.  New code should use the two-step approach:
      1. compute_threshold_elimination() → redistribute → check winner
      2. drop_lowest_candidate()         → redistribute → check winner

    Returns: (to_eliminate: list[str], actual_threshold_pct: float)
    """
    below_threshold, actual_pct = compute_threshold_elimination(vote_counts, threshold)

    # Always eliminate at least the lowest
    sorted_cands = sorted(vote_counts.items(), key=lambda x: x[1])
    lowest_num = sorted_cands[0][0]
    if lowest_num not in below_threshold:
        below_threshold.append(lowest_num)

    return below_threshold, actual_pct

# =============================================================================
# RUNOFF BALLOT GENERATION & TABULATION
# =============================================================================
#
# Triggered in Distribution Round 2+ when two or more candidates share the
# exact same lowest vote total and one must be dropped.
#
# Step 1 — Tiebreaker runoff ballot (the tied candidates only)
#           "Which candidate do you approve?  Select one."
#           Loser is eliminated; winner advances.
#
# Step 2 — Head-to-head ballot (if exactly 2 candidates remain after step 1)
#           Same question, now between the tiebreaker winner and the leader.
#           Winner of this ballot is elected to the seat.
#
# Both ballots use ballot.html with ?type=runoff in the URL.
# Results arrive as a separate Google Sheets CSV export (Runoff tab).

_ballot_base_url = ""   # set once per session when first runoff is needed

def _get_ballot_base_url():
    global _ballot_base_url
    if not _ballot_base_url:
        print(c("\n  A runoff ballot needs to be generated.", Color.CYAN))
        print(c("  Enter the GitHub Pages URL where ballot.html is hosted,", Color.GREY))
        print(c("  e.g.  https://yourname.github.io/ballot.html", Color.GREY))
        url = input("\n  Ballot URL: ").strip().rstrip("/")
        _ballot_base_url = url
    return _ballot_base_url

def _make_runoff_url(base_url, runoff_election_key, candidate_map_subset):
    """Build a ballot.html URL for a two-candidate runoff (type=runoff)."""
    cands_param = "-".join(
        f"{num}-{name.replace(' ', '%20')}"
        for num, name in candidate_map_subset.items()
    )
    return (f"{base_url}?election={runoff_election_key}"
            f"&candidates={cands_param}&seats=1&type=runoff")

def _show_runoff_qr(url, label):
    """Print the runoff URL and show a QR code if the qrcode library is available."""
    print(c(f"\n  ┌─ {label} ─────────────────────────────────────────────", Color.MAGENTA))
    print(c(f"  │  URL: {url}", Color.CYAN))
    print(c(f"  └────────────────────────────────────────────────────────", Color.MAGENTA))
    log(f"[RUNOFF URL — {label}]: {url}", also_print=False)
    try:
        import qrcode as _qr
        qr = _qr.QRCode(box_size=2, border=2)
        qr.add_data(url)
        qr.make(fit=True)
        qr.print_ascii(invert=True)
    except ImportError:
        print(c("  (Install qrcode[pil] to display QR code here, or use "
                "district3_setup.py to generate one.)", Color.GREY))

def _load_runoff_csv(filepath):
    """
    Load runoff ballot results from Google Sheets CSV export.
    Expected columns: Timestamp | Delegate Number | Election |
                      Is Test | Is Auto-Submit | Choice
    'Choice' contains the candidate LETTER the delegate selected.
    Returns ({num_str: count_int}, warnings_list).
    De-duplicates by delegate number (last submission wins).
    """
    by_delegate = {}
    warnings    = []
    try:
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader     = csv.DictReader(f)
            fieldnames = [fn.strip() for fn in (reader.fieldnames or [])]
            choice_col = next(
                (fn for fn in fieldnames if fn.strip().lower() == "choice"), None)
            if not choice_col:
                warnings.append("No 'Choice' column found in the runoff CSV. "
                                 "Check that the Apps Script wrote the correct header.")
                return {}, warnings
            for row in reader:
                if str(row.get("Is Test", "NO")).strip().upper() == "YES":
                    continue
                delegate = str(row.get("Delegate Number", "")).strip()
                choice   = str(row.get(choice_col, "")).strip()
                if delegate and choice:
                    by_delegate[delegate] = choice
    except FileNotFoundError:
        warnings.append(f"File not found: {filepath}")
        return {}, warnings

    vote_counts = {}
    for choice in by_delegate.values():
        vote_counts[choice] = vote_counts.get(choice, 0) + 1
    return vote_counts, warnings

def run_runoff(runoff_type, candidate_map_subset, election_key,
               election_label, seat_num, seat_round_label, rounds_data_out):
    """
    Orchestrate one runoff ballot event.

    runoff_type         : "tiebreaker" | "head-to-head"
    candidate_map_subset: {num_str: name_str}  — exactly 2 entries
    seat_round_label    : string used in Excel tab, e.g. "Tiebreaker" or "H2H"
    rounds_data_out     : list to append the result record to

    Returns winner_num (str) or None on failure.
    """
    nums  = list(candidate_map_subset.keys())
    names = [candidate_map_subset[n] for n in nums]

    section_header(f"RUNOFF — {election_label} — Seat {seat_num} — "
                   f"{runoff_type.replace('-',' ').title()}")

    if runoff_type == "tiebreaker":
        print(c(f"\n  Tie for lowest between: {names[0]}  and  {names[1]}", Color.MAGENTA + Color.BOLD))
        print(c("  A tiebreaker runoff ballot will be issued. Delegates select ONE candidate.", Color.GREY))
    else:
        print(c(f"\n  Head-to-head ballot: {names[0]}  vs  {names[1]}", Color.MAGENTA + Color.BOLD))
        print(c("  Delegates select ONE candidate. The winner is elected to this seat.", Color.GREY))

    # Generate and display URL + QR code
    runoff_key = f"runoff-{election_key}-{''.join(nums)}"
    base_url   = _get_ballot_base_url()
    url        = _make_runoff_url(base_url, runoff_key, candidate_map_subset)
    _show_runoff_qr(url, f"Seat {seat_num} — {runoff_type.title()}")

    print(c("\n  ─── Chair instructions ───────────────────────────────────────", Color.GREY))
    print(c("  1. Project the QR code or read the URL aloud.", Color.GREY))
    print(c("  2. Announce to delegates:", Color.WHITE))
    print(c(f'     "Which candidate do you approve — {names[0]} or {names[1]}?'
            f' You may select only one."', Color.BOLD))
    print(c("  3. Allow time to vote, then close the ballot window.", Color.GREY))
    print(c("  4. Export the Runoff tab from Google Sheets as CSV.", Color.GREY))
    print(c("  ─────────────────────────────────────────────────────────────", Color.GREY))

    input(c("\n  Press Enter when ready to load the runoff CSV...", Color.YELLOW))
    path = get_csv_filepath(f"Runoff CSV — {runoff_type}")

    vote_counts = {}
    total       = 0
    winner_num  = None

    if path:
        vote_counts, warnings = _load_runoff_csv(path)
        for w in warnings:
            print(c(f"  ⚠  {w}", Color.YELLOW))

        if vote_counts:
            total = sum(vote_counts.values())
            print(c(f"\n  Results ({total} votes cast):", Color.BOLD))
            for num in sorted(nums, key=lambda n: -vote_counts.get(n, 0)):
                name = candidate_map_subset[num]
                cnt  = vote_counts.get(num, 0)
                pct  = cnt / total * 100 if total else 0
                bar  = "█" * int(pct / 5)
                print(c(f"    {name:<28} {cnt:>4}  ({pct:>5.1f}%)  {bar}", Color.WHITE))
                log(f"  [RUNOFF {runoff_type.upper()}] {name}: {cnt} ({pct:.1f}%)",
                    also_print=False)

            max_votes      = max(vote_counts.values())
            tied_in_runoff = [n for n, v in vote_counts.items() if v == max_votes]

            if len(tied_in_runoff) == 1:
                winner_num = tied_in_runoff[0]
            else:
                print(c("\n  ⚠  Runoff itself is tied. Chair must determine the winner.", Color.RED))
                log("[RUNOFF TIE] Runoff result tied — Chair determination required.", also_print=False)
                print(c("  Enter the winner:", Color.YELLOW))
                for i, n in enumerate(tied_in_runoff, 1):
                    print(f"    {i}. {candidate_map_subset[n]}")
                while True:
                    try:
                        ch = int(input("  Enter number: ").strip())
                        if 1 <= ch <= len(tied_in_runoff):
                            winner_num = tied_in_runoff[ch - 1]
                            break
                    except ValueError:
                        pass
                    print(c("  Please enter a valid number.", Color.RED))
    else:
        # No CSV — manual entry fallback
        print(c("\n  No CSV loaded. Enter the winner manually:", Color.YELLOW))
        for i, n in enumerate(nums, 1):
            print(f"    {i}. {candidate_map_subset[n]}")
        while True:
            try:
                ch = int(input("  Enter number: ").strip())
                if 1 <= ch <= len(nums):
                    winner_num = nums[ch - 1]
                    break
            except ValueError:
                pass
            print(c("  Please enter a valid number.", Color.RED))
        vote_counts = {n: 0 for n in nums}

    if winner_num:
        loser_nums  = [n for n in nums if n != winner_num]
        winner_name = candidate_map_subset[winner_num]
        loser_names = [candidate_map_subset[n] for n in loser_nums]

        if runoff_type == "head-to-head":
            print(c(f"\n  ✓ {winner_name} ELECTED (Seat {seat_num} — head-to-head runoff)",
                    Color.GREEN + Color.BOLD))
        else:
            print(c(f"\n  {winner_name} advances.  "
                    f"{', '.join(loser_names)} eliminated.", Color.CYAN))

        log(f"[RUNOFF {runoff_type.upper()} — Seat {seat_num}] "
            f"Winner: {winner_name}  Eliminated: {loser_names}", also_print=False)

        rounds_data_out.append({
            "seat_num":        seat_num,
            "seat_round_num":  seat_round_label,
            "vote_counts":     {n: vote_counts.get(n, 0) for n in nums},
            "total_active":    total,
            "exhausted":       0,
            "majority_needed": total // 2 + 1 if total else 1,
            "elected":         [winner_num] if runoff_type == "head-to-head" else [],
            "eliminated":      loser_nums,
            "elim_reason":     f"{runoff_type} runoff ballot",
        })

    return winner_num

# =============================================================================
# ELECTION A — STATE CENTRAL COMMITTEE (SCC)   Article VI / V.K
# =============================================================================

def run_irv_one_seat(ballots, candidate_map, active_nums, election_label,
                     seat_num, max_wins_this_seat, rounds_data_out,
                     election_key=""):
    """
    Run IRV for a single seat among the given active_nums.
    Appends round data dicts to rounds_data_out (one per redistribution round).
    Each dict includes 'seat_round_num' (1, 2, 3… within this seat).
    Returns winner_num (str) or None.
    """
    active       = list(active_nums)
    seat_round   = 1   # redistribution round counter, resets to 1 for each seat

    # Show the seat ballot header once before redistribution begins
    seat_ballot_header(seat_num, election_label)

    while active:
        distribution_round_header(seat_round, seat_num, election_label)

        vote_counts, total_active, exhausted = count_round_votes(ballots, active)
        majority_needed = int(total_active * MAJORITY_THRESHOLD) + 1

        # Potential winners (> 50%)
        potential_winners = [num for num, v in vote_counts.items()
                             if v >= majority_needed]

        # Cap winners for the first seat's first redistribution round
        if seat_round == 1 and seat_num == 1:
            potential_winners = potential_winners[:max_wins_this_seat]

        winner_num = potential_winners[0] if potential_winners else None

        if winner_num:
            display_round_table(candidate_map, vote_counts, total_active, majority_needed,
                                elected=[winner_num])
            rounds_data_out.append({
                "seat_num":      seat_num,
                "seat_round_num": seat_round,
                "vote_counts":   dict(vote_counts),
                "total_active":  total_active,
                "exhausted":     exhausted,
                "majority_needed": majority_needed,
                "elected":       [winner_num],
                "eliminated":    [],
            })
            name = candidate_map.get(winner_num, winner_num)
            print(c(f"\n  ✓ {name} ELECTED (Seat {seat_num})", Color.GREEN + Color.BOLD))
            log(f"[SEAT {seat_num} — Round {seat_round}] WINNER: {name}", also_print=False)
            return winner_num

        # No winner — determine eliminations
        # Round 1: TWO-STEP process:
        #   Step A — apply 15% threshold (may eliminate multiple), redistribute
        #   Step B — check for 50%+1 winner; if none, drop lowest vote-getter
        # Round 2+: only the single lowest vote-getter is dropped
        sorted_by_votes = sorted(vote_counts.items(), key=lambda x: x[1])

        if seat_round == 1:
            # ── Step A: 15% threshold elimination ──────────────────────────
            threshold_elim, eff_pct = compute_threshold_elimination(vote_counts)

            if threshold_elim:
                elim_reason = f"below {eff_pct:.1f}% elimination threshold"
                display_round_table(candidate_map, vote_counts, total_active,
                                    majority_needed, eliminated=threshold_elim)
                rounds_data_out.append({
                    "seat_num":        seat_num,
                    "seat_round_num":  seat_round,
                    "vote_counts":     dict(vote_counts),
                    "total_active":    total_active,
                    "exhausted":       exhausted,
                    "majority_needed": majority_needed,
                    "elected":         [],
                    "eliminated":      list(threshold_elim),
                    "elim_reason":     elim_reason,
                })
                elim_names = [candidate_map.get(n, n) for n in threshold_elim]
                log(f"[SEAT {seat_num} — Round {seat_round} Step A] "
                    f"Threshold eliminated: {elim_names} ({elim_reason})",
                    also_print=False)
                for num in threshold_elim:
                    name = candidate_map.get(num, num)
                    active.remove(num)
                    print(c(f"  ✗ {name} eliminated  ({elim_reason})", Color.RED))

                if len(active) == 1:
                    winner_num = active[0]
                    name = candidate_map.get(winner_num, winner_num)
                    print(c(f"\n  Only one candidate remaining — {name} is elected.",
                            Color.YELLOW))
                    print(c(f"  ✓ {name} ELECTED (Seat {seat_num})",
                            Color.GREEN + Color.BOLD))
                    log(f"[SEAT {seat_num}] Last candidate standing: {name}",
                        also_print=False)
                    rounds_data_out.append({
                        "seat_num":       seat_num,
                        "seat_round_num": seat_round + 1,
                        "vote_counts":    {winner_num: total_active},
                        "total_active":   total_active,
                        "exhausted":      0,
                        "majority_needed": majority_needed,
                        "elected":        [winner_num],
                        "eliminated":     [],
                    })
                    return winner_num

                # Recount after threshold redistribution
                vote_counts, total_active, exhausted = count_round_votes(
                    ballots, active)
                majority_needed = int(total_active * MAJORITY_THRESHOLD) + 1

                # Check for 50%+1 winner after threshold redistribution
                potential_winners = [num for num, v in vote_counts.items()
                                     if v >= majority_needed]
                if potential_winners:
                    winner_num = potential_winners[0]
                    display_round_table(candidate_map, vote_counts, total_active,
                                        majority_needed, elected=[winner_num])
                    rounds_data_out.append({
                        "seat_num":        seat_num,
                        "seat_round_num":  seat_round,
                        "vote_counts":     dict(vote_counts),
                        "total_active":    total_active,
                        "exhausted":       exhausted,
                        "majority_needed": majority_needed,
                        "elected":         [winner_num],
                        "eliminated":      [],
                    })
                    name = candidate_map.get(winner_num, winner_num)
                    print(c(f"\n  ✓ {name} ELECTED (Seat {seat_num}) "
                            f"— winner after threshold redistribution",
                            Color.GREEN + Color.BOLD))
                    log(f"[SEAT {seat_num} — Round {seat_round} Step B] "
                        f"WINNER after threshold redistribution: {name}",
                        also_print=False)
                    return winner_num

            # ── Step B: No winner yet — drop the lowest vote-getter ────────
            # (If no threshold candidates were eliminated, we still need
            #  to drop someone to advance the election.)
            tied_lowest = drop_lowest_candidate(vote_counts)

            if len(tied_lowest) == 1:
                to_eliminate = tied_lowest
                lowest_votes = vote_counts[tied_lowest[0]]
                lowest_pct = lowest_votes / total_active * 100 if total_active > 0 else 0
                elim_reason = f"lowest vote-getter ({lowest_pct:.1f}%) — Round 1 drop"
            else:
                # Tie for lowest after threshold — use runoff ballot
                tied_names = [candidate_map.get(n, n) for n in tied_lowest]
                print(c(f"\n  Tie for lowest ({', '.join(tied_names)}) "
                        f"— runoff ballot required.", Color.YELLOW + Color.BOLD))
                log(f"[SEAT {seat_num} — Round {seat_round}] "
                    f"TIE for lowest after threshold: {tied_names}",
                    also_print=False)

                display_round_table(candidate_map, vote_counts, total_active,
                                    majority_needed, eliminated=tied_lowest)
                rounds_data_out.append({
                    "seat_num":        seat_num,
                    "seat_round_num":  seat_round,
                    "vote_counts":     dict(vote_counts),
                    "total_active":    total_active,
                    "exhausted":       exhausted,
                    "majority_needed": majority_needed,
                    "elected":         [],
                    "eliminated":      list(tied_lowest),
                    "elim_reason":     "tie for lowest — tiebreaker runoff ballot issued",
                })

                tied_map  = {n: candidate_map[n] for n in tied_lowest}
                tb_winner = run_runoff(
                    "tiebreaker", tied_map, election_key, election_label,
                    seat_num, "Tiebreaker", rounds_data_out,
                )
                if tb_winner is None:
                    return None

                to_elim_tb = [n for n in tied_lowest if n != tb_winner]
                for num in to_elim_tb:
                    active.remove(num)
                    print(c(f"  ✗ {candidate_map.get(num, num)} eliminated  "
                            f"(tiebreaker runoff)", Color.RED))
                    log(f"[SEAT {seat_num}] Tiebreaker loser: "
                        f"{candidate_map.get(num, num)}", also_print=False)

                if len(active) == 2:
                    h2h_map    = {n: candidate_map[n] for n in active}
                    h2h_winner = run_runoff(
                        "head-to-head", h2h_map, election_key, election_label,
                        seat_num, "H2H", rounds_data_out,
                    )
                    if h2h_winner:
                        return h2h_winner

                seat_round += 1
                continue

        else:
            # Round 2+: check for a tie before eliminating
            lowest_votes    = sorted_by_votes[0][1]
            tied_for_lowest = [num for num, v in vote_counts.items() if v == lowest_votes]

            if len(tied_for_lowest) == 1:
                # Clear lowest — normal single-candidate elimination
                to_eliminate = tied_for_lowest
                lowest_pct   = lowest_votes / total_active * 100 if total_active > 0 else 0
                elim_reason  = f"lowest vote-getter ({lowest_pct:.1f}%)"

            else:
                # ── Tie for lowest — runoff ballot required ─────────────────
                tied_names = [candidate_map.get(n, n) for n in tied_for_lowest]
                print(c(f"\n  Tie for lowest ({', '.join(tied_names)}) "
                        f"— runoff ballot required.", Color.YELLOW + Color.BOLD))
                log(f"[SEAT {seat_num} — Round {seat_round}] "
                    f"TIE for lowest: {tied_names}", also_print=False)

                # Show this round's vote table with all tied-lowest highlighted
                display_round_table(candidate_map, vote_counts, total_active, majority_needed,
                                    eliminated=tied_for_lowest)
                rounds_data_out.append({
                    "seat_num":        seat_num,
                    "seat_round_num":  seat_round,
                    "vote_counts":     dict(vote_counts),
                    "total_active":    total_active,
                    "exhausted":       exhausted,
                    "majority_needed": majority_needed,
                    "elected":         [],
                    "eliminated":      list(tied_for_lowest),
                    "elim_reason":     "tie for lowest — tiebreaker runoff ballot issued",
                })

                # Step 1: Tiebreaker runoff among the tied candidates
                tied_map  = {n: candidate_map[n] for n in tied_for_lowest}
                tb_winner = run_runoff(
                    "tiebreaker", tied_map, election_key, election_label,
                    seat_num, "Tiebreaker", rounds_data_out,
                )
                if tb_winner is None:
                    return None   # operator cancelled — abort this seat

                # Remove tiebreaker loser(s) from the active pool
                to_elim_tb = [n for n in tied_for_lowest if n != tb_winner]
                for num in to_elim_tb:
                    active.remove(num)
                    print(c(f"  ✗ {candidate_map.get(num, num)} eliminated  "
                            f"(tiebreaker runoff)", Color.RED))
                    log(f"[SEAT {seat_num}] Tiebreaker loser: "
                        f"{candidate_map.get(num, num)}", also_print=False)

                # Step 2: If exactly 2 candidates now remain, run head-to-head ballot
                if len(active) == 2:
                    h2h_map    = {n: candidate_map[n] for n in active}
                    h2h_winner = run_runoff(
                        "head-to-head", h2h_map, election_key, election_label,
                        seat_num, "H2H", rounds_data_out,
                    )
                    if h2h_winner:
                        return h2h_winner

                # More than 2 remain (or head-to-head failed) — continue IRV
                seat_round += 1
                continue   # skip the normal display/append/eliminate block below

        display_round_table(candidate_map, vote_counts, total_active, majority_needed,
                            eliminated=to_eliminate)

        rounds_data_out.append({
            "seat_num":        seat_num,
            "seat_round_num":  seat_round,
            "vote_counts":     dict(vote_counts),
            "total_active":    total_active,
            "exhausted":       exhausted,
            "majority_needed": majority_needed,
            "elected":         [],
            "eliminated":      list(to_eliminate),
            "elim_reason":     elim_reason,
        })
        elim_names = [candidate_map.get(n, n) for n in to_eliminate]
        log(f"[SEAT {seat_num} — Round {seat_round}] No winner.  "
            f"Eliminated: {elim_names} ({elim_reason})", also_print=False)

        for num in to_eliminate:
            name = candidate_map.get(num, num)
            active.remove(num)
            print(c(f"  ✗ {name} eliminated  ({elim_reason})", Color.RED))

        if len(active) == 1:
            # Last candidate standing — elect them
            winner_num = active[0]
            name = candidate_map.get(winner_num, winner_num)
            print(c(f"\n  Only one candidate remaining — {name} is elected.", Color.YELLOW))
            print(c(f"  ✓ {name} ELECTED (Seat {seat_num})", Color.GREEN + Color.BOLD))
            log(f"[SEAT {seat_num}] Last candidate standing: {name}", also_print=False)
            rounds_data_out.append({
                "seat_num":       seat_num,
                "seat_round_num": seat_round + 1,
                "vote_counts":    {winner_num: total_active},
                "total_active":   total_active,
                "exhausted":      0,
                "majority_needed": majority_needed,
                "elected":        [winner_num],
                "eliminated":     [],
            })
            return winner_num

        seat_round += 1

    return None

def run_scc_election(gender_label, seats, ballots, candidate_map,
                     spoiled_surrendered, output_dir):
    """
    Run one gender category of the SCC election (sequential per-seat IRV).

    SEAT RESTART LOGIC (confirmed with IDP contact 2026-04-28):
      After each winner, only the winner is removed from the pool.
      All other candidates return — eliminations do NOT carry forward.
    """
    section_header(f"SCC — {gender_label.upper()} ({seats} seats)")
    log(f"\n[ELECTION] SCC — {gender_label} — {seats} seats", also_print=False)
    log(f"[DATE] {datetime.datetime.now().isoformat()}", also_print=False)

    all_nums       = list(candidate_map.keys())
    elected_all    = []       # candidate nums elected so far (in order)
    rounds_data    = []       # all round records across all seats

    # election_key is used when generating runoff ballot URLs
    election_key = "scc-w" if "women" in gender_label.lower() else "scc-mnb"

    # Seat 1 ballot cap: no more than floor(seats * 0.5) can be elected
    # in the very first redistribution round of Seat 1.
    max_seat1_round1_wins = max(1, int(seats * SCC_FIRST_BALLOT_MAX))
    print(c(f"\n  Seat 1 Ballot rule: no more than {max_seat1_round1_wins} of "
            f"{seats} seats may be filled in the first distribution round.", Color.CYAN))
    log(f"[RULE] Seat 1 first-distribution-round max wins: {max_seat1_round1_wins}",
        also_print=False)

    # Ratification path: nominees == seats
    if len(all_nums) == seats:
        print(c(f"\n  {len(all_nums)} nominees for {seats} seats — "
                f"proceeding to ratification ballot.", Color.YELLOW))
        vote_counts, total_active, exhausted = count_round_votes(ballots, all_nums)
        majority_needed = int(total_active * MAJORITY_THRESHOLD) + 1
        elected = [num for num, v in vote_counts.items() if v >= majority_needed]
        display_round_table(candidate_map, vote_counts, total_active, majority_needed,
                            elected=elected)
        rounds_data.append({
            "round_num": 1, "seat_num": 0,
            "vote_counts": vote_counts, "total_active": total_active,
            "exhausted": exhausted, "majority_needed": majority_needed,
            "elected": elected, "eliminated": [],
        })
        log(f"[RESULT] Ratification — Elected: {[candidate_map.get(n,n) for n in elected]}", also_print=False)
        if rounds_data:
            _save_scc_excel(gender_label, candidate_map, rounds_data, output_dir)
        return elected, rounds_data, spoiled_surrendered

    # Sequential per-seat IRV
    for seat in range(1, seats + 1):
        if not candidate_map:
            print(c("\n  No candidates remaining.", Color.YELLOW))
            break

        # SEAT RESTART: pool = all original candidates minus elected winners only.
        # Eliminated candidates return to the pool for each new seat.
        # (Confirmed with IDP contact 2026-04-28)
        active_pool = [num for num in all_nums
                       if num not in elected_all]

        if len(active_pool) == 0:
            print(c(f"\n  No remaining candidates for Seat {seat}.", Color.YELLOW))
            break
        if len(active_pool) == 1:
            # Only one left — still run the count for the official record
            winner_num = active_pool[0]
            name = candidate_map.get(winner_num, winner_num)

            seat_ballot_header(seat, f"SCC {gender_label}")
            distribution_round_header(1, seat, f"SCC {gender_label}")

            vote_counts, total_active, exhausted = count_round_votes(
                ballots, active_pool)
            majority_needed = int(total_active * MAJORITY_THRESHOLD) + 1
            winner_votes = vote_counts.get(winner_num, 0)
            winner_pct = winner_votes / total_active * 100 if total_active > 0 else 0

            display_round_table(candidate_map, vote_counts, total_active,
                                majority_needed, elected=[winner_num])
            rounds_data.append({
                "seat_num":        seat,
                "seat_round_num":  1,
                "vote_counts":     dict(vote_counts),
                "total_active":    total_active,
                "exhausted":       exhausted,
                "majority_needed": majority_needed,
                "elected":         [winner_num],
                "eliminated":      [],
                "elim_reason":     "uncontested — no candidates dropped",
            })

            elected_all.append(winner_num)
            print(c(f"\n  Uncontested — no candidates dropped for low vote count.",
                    Color.YELLOW))
            print(c(f"  {name} received {winner_votes} votes ({winner_pct:.1f}% of active ballots).",
                    Color.CYAN))
            print(c(f"  ✓ {name} ELECTED (Seat {seat})",
                    Color.GREEN + Color.BOLD))
            log(f"[SEAT {seat}] Uncontested: {name} — {winner_votes} votes "
                f"({winner_pct:.1f}%)", also_print=False)
            continue

        prev_rounds_count = len(rounds_data)

        winner_num = run_irv_one_seat(
            ballots, candidate_map, active_pool,
            f"SCC {gender_label}", seat,
            max_seat1_round1_wins if seat == 1 else 1,
            rounds_data,
            election_key=election_key,
        )

        if winner_num:
            elected_all.append(winner_num)
            print(c(f"  (All non-elected candidates return to pool for next seat.)",
                    Color.GREY))
        else:
            print(c(f"\n  WARNING: No winner found for Seat {seat}.", Color.RED))

    # ── Final results ────────────────────────────────────────────────────────
    print()
    print(c(f"  FINAL RESULT — SCC {gender_label} ({seats} seats)", Color.BOLD + Color.WHITE))
    for i, num in enumerate(elected_all, 1):
        print(c(f"    {i}. {candidate_map.get(num, num)}", Color.GREEN + Color.BOLD))
    log(f"\n[FINAL] SCC {gender_label} — Elected: "
        f"{[candidate_map.get(n,n) for n in elected_all]}", also_print=False)

    # Print exceptions
    _print_exceptions(ballots, spoiled_surrendered, election_name=f"SCC {gender_label}")

    # Save Excel
    if rounds_data:
        _save_scc_excel(gender_label, candidate_map, rounds_data, output_dir)

    return elected_all, rounds_data, spoiled_surrendered

def _save_scc_excel(gender_label, candidate_map, rounds_data, output_dir):
    safe = gender_label.replace("/", "_").replace(" ", "_")
    path = os.path.join(output_dir, f"SCC_{safe}_Results.xlsx")
    create_excel_report(f"SCC — {gender_label}", candidate_map, rounds_data, path)

def run_scc(ballots_women, candidate_map_women, spoiled_women,
            ballots_mnb,   candidate_map_mnb,   spoiled_mnb,
            output_dir):
    """Run the full SCC election (Women + Men/Non-Binary)."""
    section_header("STATE CENTRAL COMMITTEE (SCC) ELECTION — Article VI")
    print(c("\n  Two separate ballots: Women (4 seats) and Men/Non-Binary (4 seats).\n", Color.CYAN))

    results = {}

    elected_w, _, _ = run_scc_election(
        "Women", 4, ballots_women, candidate_map_women, spoiled_women, output_dir)
    results["Women"] = elected_w

    if confirm("\n  Proceed to Men/Non-Binary ballot?"):
        elected_mnb, _, _ = run_scc_election(
            "Men_NonBinary", 4, ballots_mnb, candidate_map_mnb, spoiled_mnb, output_dir)
        results["Men_NonBinary"] = elected_mnb

    return results

# =============================================================================
# ELECTION B — DEI COMMITTEE CHAIR   Article IX / V.K
# =============================================================================

def run_dei_election(ballots, candidate_map, spoiled_surrendered, output_dir):
    """
    Run the DEI Committee Chair election (1 seat, IRV with special tie rule).

    Tie rule (Convention Rules V.K):
      If two or more candidates are tied for the lowest total, eliminate
      all tied candidates UNLESS doing so would leave fewer than 2 —
      in that case, hold a special tiebreak round among only the tied candidates.
    """
    section_header("DEI COMMITTEE CHAIR ELECTION — Article IX")
    log("\n[ELECTION] DEI Committee Chair — 1 seat", also_print=False)
    log(f"[DATE] {datetime.datetime.now().isoformat()}", also_print=False)

    all_nums    = list(candidate_map.keys())
    active      = list(all_nums)
    rounds_data = []
    seat_round  = 1    # redistribution round within this single-seat election
    winner_num  = None

    # DEI has one seat — show the seat ballot header once
    seat_ballot_header(1, "DEI Committee Chair")

    while not winner_num and active:
        distribution_round_header(seat_round, 1, "DEI Committee Chair")

        vote_counts, total_active, exhausted = count_round_votes(ballots, active)
        majority_needed = int(total_active * MAJORITY_THRESHOLD) + 1

        # Check for winner
        best_num = max(vote_counts, key=vote_counts.get)
        if vote_counts[best_num] >= majority_needed:
            winner_num = best_num
            display_round_table(candidate_map, vote_counts, total_active, majority_needed,
                                elected=[winner_num])
            rounds_data.append({
                "seat_num":       1,
                "seat_round_num": seat_round,
                "vote_counts":    dict(vote_counts), "total_active": total_active,
                "exhausted":      exhausted, "majority_needed": majority_needed,
                "elected":        [winner_num], "eliminated": [],
            })
            name = candidate_map.get(winner_num, winner_num)
            print(c(f"\n  ✓ {name} ELECTED as DEI Committee Chair", Color.GREEN + Color.BOLD))
            log(f"[SEAT 1 — Round {seat_round}] WINNER: {name}", also_print=False)
            break

        # No winner — apply elimination rule:
        #   Distribution Round 1: TWO-STEP process:
        #     Step A — 15% threshold (may eliminate multiple), redistribute
        #     Step B — check for 50%+1 winner; if none, drop lowest
        #   Distribution Round 2+: only the single lowest vote-getter drops
        # Tie rule (both rounds): if multiple candidates share the lowest total and
        # eliminating all would leave < 2, hold a special tiebreak runoff ballot.
        sorted_cands = sorted(vote_counts.items(), key=lambda x: x[1])
        to_eliminate = []
        elim_reason  = ""

        if seat_round == 1:
            # ── Step A: 15% threshold elimination ──────────────────────────
            threshold_elim, eff_pct = compute_threshold_elimination(vote_counts)

            # Guarantee ≥ 2 survivors (DEI-specific safeguard)
            survivors = [n for n in active if n not in threshold_elim]
            while len(survivors) < 2 and len(threshold_elim) > 1:
                by_votes = sorted([(n, vote_counts[n]) for n in threshold_elim],
                                  key=lambda x: x[1])
                threshold_elim = [n for n, _ in by_votes[:-1]]
                survivors = [n for n in active if n not in threshold_elim]

            if threshold_elim:
                elim_reason = f"below {eff_pct:.1f}% elimination threshold"
                display_round_table(candidate_map, vote_counts, total_active,
                                    majority_needed, eliminated=threshold_elim)
                rounds_data.append({
                    "seat_num": 1, "seat_round_num": seat_round,
                    "vote_counts": dict(vote_counts), "total_active": total_active,
                    "exhausted": exhausted, "majority_needed": majority_needed,
                    "elected": [], "eliminated": list(threshold_elim),
                    "elim_reason": elim_reason,
                })
                elim_names = [candidate_map.get(n, n) for n in threshold_elim]
                log(f"[SEAT 1 — Round {seat_round} Step A] "
                    f"Threshold eliminated: {elim_names} ({elim_reason})",
                    also_print=False)
                for num in threshold_elim:
                    name = candidate_map.get(num, num)
                    active.remove(num)
                    print(c(f"  ✗ {name} eliminated  ({elim_reason})", Color.RED))

                if len(active) == 1:
                    winner_num = active[0]
                    name = candidate_map.get(winner_num, winner_num)
                    print(c(f"\n  Only one candidate remaining — {name} is elected.",
                            Color.YELLOW))
                    print(c(f"  ✓ {name} ELECTED as DEI Committee Chair",
                            Color.GREEN + Color.BOLD))
                    log(f"[SEAT 1] Last remaining: {name}", also_print=False)
                    seat_round += 1
                    continue   # will exit while loop since winner_num is set

                # Recount after threshold redistribution
                vote_counts, total_active, exhausted = count_round_votes(
                    ballots, active)
                majority_needed = int(total_active * MAJORITY_THRESHOLD) + 1

                # Check for 50%+1 winner after threshold redistribution
                best_num = max(vote_counts, key=vote_counts.get)
                if vote_counts[best_num] >= majority_needed:
                    winner_num = best_num
                    display_round_table(candidate_map, vote_counts, total_active,
                                        majority_needed, elected=[winner_num])
                    rounds_data.append({
                        "seat_num": 1, "seat_round_num": seat_round,
                        "vote_counts": dict(vote_counts), "total_active": total_active,
                        "exhausted": exhausted, "majority_needed": majority_needed,
                        "elected": [winner_num], "eliminated": [],
                    })
                    name = candidate_map.get(winner_num, winner_num)
                    print(c(f"\n  ✓ {name} ELECTED as DEI Committee Chair "
                            f"— winner after threshold redistribution",
                            Color.GREEN + Color.BOLD))
                    log(f"[SEAT 1 — Round {seat_round} Step B] "
                        f"WINNER after threshold: {name}", also_print=False)
                    break

            # ── Step B: No winner yet — drop the lowest vote-getter ────────
            tied_lowest = drop_lowest_candidate(vote_counts)

            # DEI safeguard: if eliminating all tied would leave < 2,
            # hold a special tiebreak runoff ballot (Convention Rules V.K)
            if len(tied_lowest) > 1:
                after_elim = len(active) - len(tied_lowest)
                if after_elim >= 2:
                    to_eliminate = tied_lowest
                    names = [candidate_map.get(n, n) for n in tied_lowest]
                    print(c(f"\n  Tie for lowest ({', '.join(names)}) — all eliminated.",
                            Color.YELLOW))
                    elim_reason = "tied for lowest vote-getter"
                else:
                    # Tiebreak runoff ballot (NOT a recount of existing rankings)
                    names = [candidate_map.get(n, n) for n in tied_lowest]
                    print(c(f"\n  Tie for lowest: eliminating all would leave < 2.",
                            Color.YELLOW + Color.BOLD))
                    print(c(f"  Special tiebreak runoff ballot among: "
                            f"{', '.join(names)}", Color.YELLOW))
                    log(f"[SEAT 1 — Round {seat_round}] TIEBREAK RUNOFF: {names}",
                        also_print=False)

                    display_round_table(candidate_map, vote_counts, total_active,
                                        majority_needed, eliminated=tied_lowest)
                    rounds_data.append({
                        "seat_num": 1, "seat_round_num": seat_round,
                        "vote_counts": dict(vote_counts), "total_active": total_active,
                        "exhausted": exhausted, "majority_needed": majority_needed,
                        "elected": [], "eliminated": list(tied_lowest),
                        "elim_reason": "tie for lowest — tiebreaker runoff ballot issued",
                    })

                    tied_map = {n: candidate_map[n] for n in tied_lowest}
                    tb_winner = run_runoff(
                        "tiebreaker", tied_map, "dei-chair", "DEI Committee Chair",
                        1, "Tiebreaker", rounds_data,
                    )
                    if tb_winner is None:
                        break   # operator cancelled

                    to_elim_tb = [n for n in tied_lowest if n != tb_winner]
                    for num in to_elim_tb:
                        active.remove(num)
                        print(c(f"  ✗ {candidate_map.get(num, num)} eliminated  "
                                f"(tiebreaker runoff)", Color.RED))
                        log(f"[SEAT 1] Tiebreaker loser: "
                            f"{candidate_map.get(num, num)}", also_print=False)

                    if len(active) == 2:
                        h2h_map = {n: candidate_map[n] for n in active}
                        h2h_winner = run_runoff(
                            "head-to-head", h2h_map, "dei-chair",
                            "DEI Committee Chair", 1, "H2H", rounds_data,
                        )
                        if h2h_winner:
                            winner_num = h2h_winner

                    seat_round += 1
                    continue
            else:
                to_eliminate = tied_lowest
                lowest_votes = vote_counts[tied_lowest[0]]
                lowest_pct = lowest_votes / total_active * 100 if total_active > 0 else 0
                elim_reason = f"lowest vote-getter ({lowest_pct:.1f}%)"

        else:
            # ── Round 2+: lowest drops ────────────────────────────────────────
            lowest_votes = sorted_cands[0][1]
            tied_lowest  = [num for num, v in vote_counts.items() if v == lowest_votes]

            if len(tied_lowest) > 1:
                after_elim = len(active) - len(tied_lowest)
                if after_elim >= 2:
                    to_eliminate = tied_lowest
                    names = [candidate_map.get(n, n) for n in tied_lowest]
                    print(c(f"\n  Tie for lowest ({', '.join(names)}) — all eliminated.",
                            Color.YELLOW))
                    elim_reason = "tied for lowest vote-getter"
                else:
                    # Tiebreak runoff ballot (Convention Rules V.K)
                    names = [candidate_map.get(n, n) for n in tied_lowest]
                    print(c(f"\n  Tie for lowest: eliminating all would leave < 2.",
                            Color.YELLOW + Color.BOLD))
                    print(c(f"  Special tiebreak runoff ballot among: "
                            f"{', '.join(names)}", Color.YELLOW))
                    log(f"[SEAT 1 — Round {seat_round}] TIEBREAK RUNOFF: {names}",
                        also_print=False)

                    display_round_table(candidate_map, vote_counts, total_active,
                                        majority_needed, eliminated=tied_lowest)
                    rounds_data.append({
                        "seat_num": 1, "seat_round_num": seat_round,
                        "vote_counts": dict(vote_counts), "total_active": total_active,
                        "exhausted": exhausted, "majority_needed": majority_needed,
                        "elected": [], "eliminated": list(tied_lowest),
                        "elim_reason": "tie for lowest — tiebreaker runoff ballot issued",
                    })

                    tied_map = {n: candidate_map[n] for n in tied_lowest}
                    tb_winner = run_runoff(
                        "tiebreaker", tied_map, "dei-chair", "DEI Committee Chair",
                        1, "Tiebreaker", rounds_data,
                    )
                    if tb_winner is None:
                        break

                    to_elim_tb = [n for n in tied_lowest if n != tb_winner]
                    for num in to_elim_tb:
                        active.remove(num)
                        print(c(f"  ✗ {candidate_map.get(num, num)} eliminated  "
                                f"(tiebreaker runoff)", Color.RED))
                        log(f"[SEAT 1] Tiebreaker loser: "
                            f"{candidate_map.get(num, num)}", also_print=False)

                    if len(active) == 2:
                        h2h_map = {n: candidate_map[n] for n in active}
                        h2h_winner = run_runoff(
                            "head-to-head", h2h_map, "dei-chair",
                            "DEI Committee Chair", 1, "H2H", rounds_data,
                        )
                        if h2h_winner:
                            winner_num = h2h_winner

                    seat_round += 1
                    continue
            else:
                to_eliminate = [sorted_cands[0][0]]
                lowest_pct = lowest_votes / total_active * 100 if total_active > 0 else 0
                elim_reason = f"lowest vote-getter ({lowest_pct:.1f}%)"

        display_round_table(candidate_map, vote_counts, total_active, majority_needed,
                            eliminated=to_eliminate)

        rounds_data.append({
            "seat_num": 1, "seat_round_num": seat_round,
            "vote_counts": dict(vote_counts), "total_active": total_active,
            "exhausted": exhausted, "majority_needed": majority_needed,
            "elected": [], "eliminated": list(to_eliminate),
            "elim_reason": elim_reason,
        })
        log(f"[SEAT 1 — Round {seat_round}] Eliminated: "
            f"{[candidate_map.get(n,n) for n in to_eliminate]} — {elim_reason}",
            also_print=False)

        for num in to_eliminate:
            name = candidate_map.get(num, num)
            active.remove(num)
            print(c(f"  ✗ {name} eliminated  ({elim_reason})", Color.RED))

        if len(active) == 1:
            winner_num = active[0]
            name = candidate_map.get(winner_num, winner_num)
            print(c(f"\n  Only one candidate remaining — {name} is elected.",
                    Color.YELLOW))
            print(c(f"  ✓ {name} ELECTED as DEI Committee Chair",
                    Color.GREEN + Color.BOLD))
            log(f"[SEAT 1 — Round {seat_round}] Last remaining: {name}",
                also_print=False)

        seat_round += 1

    # Exceptions + Excel
    _print_exceptions(ballots, spoiled_surrendered, election_name="DEI Committee Chair")
    if rounds_data:
        path = os.path.join(output_dir, "DEI_Chair_Results.xlsx")
        create_excel_report("DEI Committee Chair", candidate_map, rounds_data, path)

    if winner_num:
        log(f"\n[FINAL] DEI Chair — Elected: {candidate_map.get(winner_num, winner_num)}",
            also_print=False)
    return winner_num

# =============================================================================
# ELECTION C — STATE CONVENTION COMMITTEE (14 seats)   Article XII
# =============================================================================
#
# Slate vote: delegates vote YES or NO on the full slate.
# If slate approved → all nominees elected.
# If slate fails → individual IRV: 14 highest vote-getters elected.
# CSV column: "Slate Vote" containing YES / NO

def run_committee_election(output_dir):
    """Run the State Convention Committee election (slate vote, 14 seats)."""
    section_header("STATE CONVENTION COMMITTEE ELECTION — Article XII")
    log("\n[ELECTION] State Convention Committee — 14 seats (slate vote)", also_print=False)
    log(f"[DATE] {datetime.datetime.now().isoformat()}", also_print=False)

    seats = 14
    print(c("\n  This is a slate vote (YES / NO on the full slate).", Color.CYAN))
    print(c("  Nominees should already be recorded by the Committee.", Color.GREY))

    # Load CSV
    choice = menu("How will slate votes be loaded?", [
        "Google Sheets CSV export",
        "Paper ballot CSV  (Slate Vote column: YES or NO)",
        "Enter totals manually",
    ])

    yes_votes = 0
    no_votes  = 0
    nominee_names = []

    if choice in (1, 2):
        label = "Convention Committee — Slate Vote"
        path  = get_csv_filepath(label)
        if path:
            try:
                with open(path, newline="", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        vote = str(row.get("Slate Vote", row.get("slate_vote", ""))).strip().upper()
                        if str(row.get("Is Test","NO")).strip().upper() == "YES":
                            continue
                        if vote in ("YES", "Y", "1"):
                            yes_votes += 1
                        elif vote in ("NO", "N", "0"):
                            no_votes  += 1
                log(f"[LOAD] Slate CSV: {path}", also_print=False)
            except FileNotFoundError:
                print(c(f"  File not found: {path}", Color.RED))
    else:
        while True:
            try:
                yes_votes = int(input("\n  Total YES votes: ").strip())
                no_votes  = int(input("  Total NO votes:  ").strip())
                break
            except ValueError:
                print(c("  Please enter whole numbers.", Color.RED))

    total = yes_votes + no_votes
    print(c(f"\n  YES: {yes_votes}   NO: {no_votes}   Total: {total}", Color.CYAN))
    log(f"\n[SLATE VOTE] YES: {yes_votes}  NO: {no_votes}  Total: {total}", also_print=False)

    majority_needed = int(total * MAJORITY_THRESHOLD) + 1
    slate_approved  = yes_votes >= majority_needed

    if slate_approved:
        print(c(f"\n  Slate APPROVED ({yes_votes}/{total} = {yes_votes/total*100:.1f}%)", Color.GREEN + Color.BOLD))
        print(c("  All nominated slate members are elected.", Color.GREEN))
        log("[RESULT] Slate approved — all nominees elected.", also_print=False)
    else:
        print(c(f"\n  Slate FAILED ({yes_votes}/{total} = {yes_votes/total*100:.1f}%)", Color.RED + Color.BOLD))
        print(c("  Per Article XII.F.2: individual vote — 14 highest vote-getters elected.", Color.YELLOW))
        log("[RESULT] Slate failed — individual vote required.", also_print=False)
        print(c("\n  Individual vote not yet supported in this version.\n"
                "  Please run the election using the aggregate method and record results.", Color.YELLOW))

    # Save simple Excel summary
    if EXCEL_AVAILABLE:
        path = os.path.join(output_dir, "Convention_Committee_Slate_Vote.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Slate Vote"
        hdr_fill = PatternFill("solid", fgColor="1F3864")
        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        headers  = ["Vote Type", "Count", "% of Total", "Result"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill;  cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for row_i, (label, count) in enumerate([("YES", yes_votes), ("NO", no_votes)], 2):
            pct  = count / total * 100 if total else 0
            vals = [label, count, f"{pct:.1f}%",
                    ("APPROVED" if slate_approved else "FAILED") if label == "YES" else ""]
            for col, val in enumerate(vals, 1):
                ws.cell(row=row_i, column=col, value=val)
        ws.column_dimensions["A"].width = 14
        for ch in ["B","C","D"]:
            ws.column_dimensions[ch].width = 18
        wb.save(path)
        print(c(f"\n  Excel saved → {path}", Color.GREEN))

    return slate_approved

# =============================================================================
# EXCEPTIONS REPORT
# =============================================================================

def _print_exceptions(ballots, spoiled_surrendered, election_name=""):
    """Print a summary of spoiled, exhausted, and surrendered credential ballots."""
    exhausted = sum(
        1 for b in ballots
        if not b.is_spoiled and b.get_first_active_choice(list(b.rankings.keys())) is None
    )
    n_spoiled     = sum(1 for b in ballots if b.is_spoiled)
    n_surrendered = len(spoiled_surrendered)

    print()
    print(c("  ─── EXCEPTIONS REPORT ───────────────────────────────────", Color.GREY))
    print(c(f"  Election : {election_name}", Color.GREY))
    print(c(f"  Spoiled / Invalid ballots   : {n_spoiled}", Color.YELLOW if n_spoiled else Color.GREY))
    print(c(f"  Surrendered Credentials     : {n_surrendered}", Color.YELLOW if n_surrendered else Color.GREY))
    print(c(f"  Exhausted (blank/no choice) : {exhausted}", Color.GREY))
    print(c("  ─────────────────────────────────────────────────────────", Color.GREY))
    print()

    log(f"\n[EXCEPTIONS — {election_name}]", also_print=False)
    log(f"  Spoiled/Invalid   : {n_spoiled}", also_print=False)
    log(f"  Surrendered Creds : {n_surrendered}", also_print=False)
    log(f"  Exhausted         : {exhausted}", also_print=False)
    for b in spoiled_surrendered:
        log(f"  SURRENDERED: Delegate {b.delegate_number} — {b.spoil_reason}", also_print=False)

# =============================================================================
# EXCEL EXPORT
# =============================================================================

def add_redistribution_flow_tab(wb, election_name, candidate_map, rounds_data):
    """
    Appends a "Redistribution Flow" worksheet (inserted at position 0 so it
    opens first) to the workbook.

    Layout — one section per seat:
      • A seat header row (dark blue, shows who is and isn't in the pool)
      • A column header row:
          Candidate | Round 1 Votes | Δ | Round 2 Votes | Δ | … | Outcome
      • One row per candidate:
          – Votes column: "N  (XX.X%)" with green fill if elected that round,
            red if eliminated, grey if already gone
          – Δ column: change from prior round (light green/red fill for ±)
          – Outcome column: "✓ ELECTED (Round N)" or "✗ Eliminated (Round N)"
      • Not-in-pool rows greyed out with "Not in this seat's pool"
      • Three summary rows: Active Ballots | Exhausted | Majority Needed
    """
    if not EXCEL_AVAILABLE:
        return

    ws = wb.create_sheet(title="Redistribution Flow", index=0)

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1F3864")  # dark navy
    seat_fill  = PatternFill("solid", fgColor="305496")  # medium navy
    elect_fill = PatternFill("solid", fgColor="C6EFCE")  # green
    elim_fill  = PatternFill("solid", fgColor="FFCCCC")  # red/pink
    na_fill    = PatternFill("solid", fgColor="EBEBEB")  # light grey
    stat_fill  = PatternFill("solid", fgColor="FFF2CC")  # yellow
    gain_fill  = PatternFill("solid", fgColor="E2EFDA")  # light green (Δ +)
    loss_fill  = PatternFill("solid", fgColor="FCE4D6")  # light red   (Δ -)

    hdr_font  = Font(name="Calibri", bold=True,  color="FFFFFF", size=11)
    seat_font = Font(name="Calibri", bold=True,  color="FFFFFF", size=12)
    bold_font = Font(name="Calibri", bold=True,  size=11)
    norm_font = Font(name="Calibri", size=11)
    grey_font = Font(name="Calibri", size=10, italic=True, color="808080")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))

    def ap(cell, fill=None, font=None, align="center", wrap=False):
        if fill:
            cell.fill = fill
        cell.font      = font or norm_font
        cell.alignment = Alignment(horizontal=align, vertical="center",
                                   wrap_text=